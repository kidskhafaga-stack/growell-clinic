"""Bulk patient import: parsing Excel/CSV and building the template.

Header matching is alias-based and case-insensitive, accepting both English
keys and common Arabic column names so a clinic's existing sheet often works
without renaming columns.
"""
import csv
import io
from datetime import date, datetime

# Canonical columns in template order: (key, required, arabic_label, sample).
IMPORT_COLUMNS = [
    ("reference_number", False, "رقم الملف المرجعي", "1043"),
    ("full_name", True, "الاسم بالعربي", "أحمد محمد"),
    ("full_name_en", False, "الاسم بالإنجليزي", "Ahmed Mohamed"),
    ("date_of_birth", True, "تاريخ الميلاد (YYYY-MM-DD)", "2022-03-01"),
    ("gender", True, "الجنس (male/female)", "male"),
    ("national_id", False, "الرقم القومي", "22203011234567"),
    ("blood_type", False, "فصيلة الدم", "O+"),
    ("allergies", False, "الحساسية", "البنسلين"),
    ("chronic_diseases", False, "الأمراض المزمنة", ""),
    ("family_name", False, "اسم الأسرة (لربط الأخوة)", "عائلة محمد"),
    ("parent_name", False, "اسم ولي الأمر", "محمد علي"),
    ("parent_relation", False, "صلة القرابة (father/mother)", "father"),
    ("parent_phone", False, "هاتف ولي الأمر", "01000000000"),
    ("parent_phone_alt", False, "هاتف ولي الأمر الإضافي", "01100000000"),
    ("parent_national_id", False, "الرقم القومي لولي الأمر", "28001011234567"),
    ("parent_email", False, "بريد ولي الأمر", "parent@mail.com"),
    ("parent_occupation", False, "مهنة ولي الأمر", "مهندس"),
    ("parent_nationality", False, "جنسية ولي الأمر", "مصري"),
    ("parent_address", False, "عنوان ولي الأمر", "الإسكندرية - سموحة"),
    ("client_category", False, "فئة العميل (normal/friend/relative/employee)", "normal"),
    ("notes", False, "ملاحظات", ""),
]

# Accepted header aliases (lower-cased) -> canonical key.
COLUMN_ALIASES = {
    "reference_number": "reference_number", "ref": "reference_number",
    "رقم الملف": "reference_number", "رقم الملف المرجعي": "reference_number",
    "الرقم المرجعي": "reference_number", "رقم قديم": "reference_number",
    "full_name": "full_name", "name": "full_name", "الاسم": "full_name",
    "الاسم بالعربي": "full_name", "اسم المريض": "full_name",
    "full_name_en": "full_name_en", "name_en": "full_name_en",
    "english name": "full_name_en", "الاسم بالإنجليزي": "full_name_en",
    "الاسم بالانجليزي": "full_name_en",
    "date_of_birth": "date_of_birth", "dob": "date_of_birth",
    "birth": "date_of_birth", "تاريخ الميلاد": "date_of_birth",
    "تاريخ الميلاد (yyyy-mm-dd)": "date_of_birth", "الميلاد": "date_of_birth",
    "gender": "gender", "sex": "gender", "الجنس": "gender", "النوع": "gender",
    "الجنس (male/female)": "gender",
    "national_id": "national_id", "nid": "national_id",
    "الرقم القومي": "national_id",
    "blood_type": "blood_type", "blood": "blood_type",
    "فصيلة الدم": "blood_type", "الفصيلة": "blood_type",
    "allergies": "allergies", "الحساسية": "allergies", "حساسية": "allergies",
    "chronic_diseases": "chronic_diseases", "chronic": "chronic_diseases",
    "الأمراض المزمنة": "chronic_diseases", "الامراض المزمنة": "chronic_diseases",
    "family_name": "family_name", "family": "family_name",
    "الأسرة": "family_name", "الاسرة": "family_name",
    "اسم الأسرة": "family_name", "اسم الاسرة": "family_name",
    "اسم الأسرة (لربط الأخوة)": "family_name", "العائلة": "family_name",
    "parent_name": "parent_name", "parent": "parent_name",
    "ولي الأمر": "parent_name", "ولي الامر": "parent_name",
    "اسم ولي الأمر": "parent_name", "اسم ولي الامر": "parent_name",
    "parent_relation": "parent_relation", "relation": "parent_relation",
    "صلة القرابة": "parent_relation", "صلة القرابة (father/mother)": "parent_relation",
    "parent_phone": "parent_phone", "phone": "parent_phone",
    "هاتف ولي الأمر": "parent_phone", "الهاتف": "parent_phone",
    "التليفون": "parent_phone", "الموبايل": "parent_phone",
    "parent_phone_alt": "parent_phone_alt", "phone2": "parent_phone_alt",
    "هاتف إضافي": "parent_phone_alt", "هاتف اضافي": "parent_phone_alt",
    "هاتف ولي الأمر الإضافي": "parent_phone_alt", "تليفون آخر": "parent_phone_alt",
    "parent_national_id": "parent_national_id",
    "الرقم القومي لولي الأمر": "parent_national_id", "قومي ولي الأمر": "parent_national_id",
    "parent_email": "parent_email", "email": "parent_email",
    "بريد ولي الأمر": "parent_email", "الايميل": "parent_email", "الإيميل": "parent_email",
    "parent_occupation": "parent_occupation", "occupation": "parent_occupation",
    "مهنة ولي الأمر": "parent_occupation", "المهنة": "parent_occupation", "الوظيفة": "parent_occupation",
    "parent_nationality": "parent_nationality", "nationality": "parent_nationality",
    "جنسية ولي الأمر": "parent_nationality", "الجنسية": "parent_nationality",
    "parent_address": "parent_address", "address": "parent_address",
    "عنوان ولي الأمر": "parent_address", "العنوان": "parent_address", "السكن": "parent_address",
    "client_category": "client_category", "category": "client_category",
    "فئة العميل": "client_category", "الفئة": "client_category",
    "فئة العميل (normal/friend/relative/employee)": "client_category",
    "notes": "notes", "ملاحظات": "notes",
}

ALLOWED_IMPORT_EXTENSIONS = {"xlsx", "csv"}

_GENDER_MAP = {
    "male": "male", "m": "male", "ذكر": "male", "ولد": "male",
    "female": "female", "f": "female", "أنثى": "female", "انثى": "female",
    "بنت": "female",
}

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%d.%m.%Y",
]


def allowed_import_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMPORT_EXTENSIONS
    )


def _norm_header(value):
    return str(value or "").strip().lower()


def map_headers(raw_headers):
    """Map a sheet's header row to canonical column keys (by position)."""
    mapping = {}
    for idx, h in enumerate(raw_headers):
        key = COLUMN_ALIASES.get(_norm_header(h))
        if key:
            mapping[idx] = key
    return mapping


# Required canonical keys (a patient cannot be created without these).
REQUIRED_KEYS = {"full_name", "date_of_birth", "gender"}


def import_fields():
    """Field metadata for the column-mapping UI: (key, required, sample)."""
    return [(key, required, sample) for key, required, _label, sample in IMPORT_COLUMNS]


def guess_mapping(headers):
    """Best-effort {canonical_key: column_index} from header aliases.

    The first column matching a key wins, so duplicate headers don't clobber
    an earlier good match.
    """
    mapping = {}
    for idx, h in enumerate(headers):
        key = COLUMN_ALIASES.get(_norm_header(h))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def build_rows(data_rows, mapping):
    """Build canonical-key dicts from raw rows using {key: column_index}.

    ``mapping`` indices that are None/out of range are treated as "ignore".
    """
    rows = []
    for raw in data_rows:
        record = {}
        for key, idx in mapping.items():
            if idx is None or idx < 0 or idx >= len(raw):
                continue
            val = raw[idx]
            if isinstance(val, str):
                val = val.strip()
            record[key] = val
        rows.append(record)
    return rows


# First-name particles that bind to the next word, so the child's own name is
# two tokens ("عبد الله", "أبو بكر") — drop both when deriving the father.
_COMPOUND_FIRST = {"عبد", "ابو", "أبو", "أم", "ام"}


def derive_guardian_name(child_name):
    """Father's name ≈ the child's name without their own (first) name.

    Arabic full names run child → father → grandfather → family, so dropping
    the child's first name yields a good default guardian name (to be verified
    later — it's flagged on import). Handles two-word first names like
    "عبد الله"; the attached spelling "عبدالله" is already a single token.
    Returns None when there isn't enough to derive.
    """
    parts = (child_name or "").strip().split()
    first_len = 2 if (parts and parts[0] in _COMPOUND_FIRST) else 1
    if len(parts) <= first_len:
        return None
    return " ".join(parts[first_len:])


# How many words of a guardian's name identify the household. Father plus
# grandfather: enough to tell two families apart, and short enough that it
# does not change when somebody types one more ancestor than the last person
# did.
GUARDIAN_KEY_WORDS = 2


def family_key(guardian_name):
    """A stable grouping key for one household, from a guardian's name.

    Grouping by the guardian name as typed is what scattered a family across
    three records. Two things break it, and both are ordinary:

    **Different lengths.** Egyptian names run child → father → grandfather →
    family, and a clinic's sheet records as many as whoever filled it knew.
    "زياد محمود سعيد أحمد" and "عمر محمود سعيد" are brothers, and yield
    "محمود سعيد أحمد" and "محمود سعيد" — two families.

    **Spelling.** "أحمد" and "احمد" are one name to every human being and two
    strings to a computer, so the same household splits on whichever keyboard
    the typist had.

    Taking a *fixed* number of leading words, folded, fixes both: the father
    and grandfather are the same in every recording of the same family,
    however much of the rest was written down.
    """
    from app.utils.history_import import normalise_arabic

    words = [w for w in normalise_arabic(guardian_name or "").split(" ") if w]
    if not words:
        return ""
    return " ".join(words[:GUARDIAN_KEY_WORDS])


def normalize_phone(value):
    """Digits-only phone for sibling matching, or None."""
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits or None


def parse_gender(value):
    return _GENDER_MAP.get(_norm_header(value))


def parse_date(value):
    """Parse a cell into a ``date`` or return None if unparseable/empty."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    # Excel may hand us a float serial as text occasionally; ignore those.
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def read_rows(file_storage):
    """Read an uploaded .xlsx/.csv into a list of canonical-key dicts.

    Returns (rows, error). ``rows`` is a list of dicts keyed by canonical
    column name with raw cell values; row numbers are 1-based data rows.
    """
    filename = (file_storage.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "xlsx":
        return _read_xlsx(file_storage)
    if ext == "csv":
        return _read_csv(file_storage)
    return [], "unsupported"


def read_matrix(file_storage):
    """Read an uploaded .xlsx/.csv into (headers, data_rows, error).

    Unlike :func:`read_rows` this does NOT require a recognised name column —
    the caller maps columns afterwards. ``headers`` is a list of header
    strings; ``data_rows`` is a list of value lists aligned to ``headers``.
    """
    filename = (file_storage.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "xlsx":
        matrix, error = _read_xlsx_matrix(file_storage)
    elif ext == "csv":
        matrix, error = _read_csv_matrix(file_storage)
    else:
        return [], [], "unsupported"

    if error:
        return [], [], error
    if not matrix:
        return [], [], "empty"

    headers = ["" if h is None else str(h).strip() for h in matrix[0]]
    width = len(headers)
    data = []
    for raw in matrix[1:]:
        if not any(str(c).strip() for c in raw if c is not None):
            continue  # skip fully blank rows
        row = list(raw)
        if len(row) < width:  # pad short rows so column indices stay aligned
            row += [None] * (width - len(row))
        data.append(row)
    if not data:
        return headers, [], "empty"
    return headers, data, None


def _rows_from_matrix(matrix):
    if not matrix:
        return [], "empty"
    headers = matrix[0]
    mapping = map_headers(headers)
    if "full_name" not in mapping.values():
        return [], "no_name_column"
    rows = []
    for raw in matrix[1:]:
        if not any(str(c).strip() for c in raw if c is not None):
            continue  # skip fully blank rows
        record = {}
        for idx, key in mapping.items():
            if idx < len(raw):
                val = raw[idx]
                record[key] = val.strip() if isinstance(val, str) else val
        rows.append(record)
    return rows, None


def _read_xlsx_matrix(file_storage):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_storage, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - surface a friendly message
        return [], "unreadable"
    ws = wb.active
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    return matrix, None


def _read_csv_matrix(file_storage):
    raw = file_storage.read()
    if isinstance(raw, bytes):
        # utf-8-sig strips a BOM that Excel-exported CSVs often carry.
        text = raw.decode("utf-8-sig", errors="replace")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    matrix = [row for row in reader]
    return matrix, None


def _read_xlsx(file_storage):
    matrix, error = _read_xlsx_matrix(file_storage)
    if error:
        return [], error
    return _rows_from_matrix(matrix)


def _read_csv(file_storage):
    matrix, error = _read_csv_matrix(file_storage)
    if error:
        return [], error
    return _rows_from_matrix(matrix)


def build_template_workbook():
    """Return a BytesIO .xlsx template with headers, a sample, and instructions."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"

    header_fill = PatternFill("solid", fgColor="198754")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [c[0] for c in IMPORT_COLUMNS]
    sample = [c[3] for c in IMPORT_COLUMNS]
    ws.append(headers)
    ws.append(sample)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Instructions sheet (Arabic descriptions).
    info = wb.create_sheet("تعليمات")
    info.append(["العمود (Column)", "الوصف", "إلزامي؟"])
    for key, required, label, _ in IMPORT_COLUMNS:
        info.append([key, label, "نعم" if required else "لا"])
    for col in ("A", "B", "C"):
        info.column_dimensions[col].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template_csv():
    """Return CSV template text (BOM included for Excel compatibility)."""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([c[0] for c in IMPORT_COLUMNS])
    writer.writerow([c[3] for c in IMPORT_COLUMNS])
    return "﻿" + out.getvalue()
