"""Patients & Families module (Phase 2).

Covers patient CRUD with manual/auto file numbers, photo upload, medical
alerts, family grouping, parents/guardians and sibling linking.
"""
import json
import os
import uuid
from datetime import date, datetime
from datetime import time as dtime

from flask import (
    Response,
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required

from app.blueprints.patients import patients_bp
from app.extensions import db
from app.i18n import t
from app.models import (
    BLOOD_TYPES,
    CLIENT_CATEGORIES,
    Family,
    GENDERS,
    CONSENT_TYPES,
    PARENT_RELATIONS,
    ActivityLog,
    Consent,
    Parent,
    Patient,
    PatientAttachment,
    PatientProblem,
)
from app.utils.clock import local_today
from app.utils.uploads import ATTACHMENT_KINDS, remove_document, save_document
from app.utils.decorators import capability_required, client_ip, module_required
from app.utils.paging import paginate
from app.utils.imports import (
    allowed_import_file,
    build_rows,
    build_template_csv,
    build_template_workbook,
    derive_guardian_name,
    guess_mapping,
    import_fields,
    normalize_phone,
    parse_date,
    parse_gender,
    read_matrix,
    REQUIRED_KEYS,
)
from app.utils.patients import (
    apply_patient_search,
    delete_patient_photo,
    generate_patient_number,
    patient_number_allocator,
    save_patient_photo,
)

MODULE = "patients"


def _upload_dir():
    return os.path.join(current_app.static_folder, "uploads", "patients")


@patients_bp.route("/<int:patient_id>/documents", methods=["POST"])
@module_required(MODULE)
@capability_required("patient_medical")
def upload_document(patient_id):
    """Upload a document (lab/imaging/report) straight to the patient file."""
    patient = db.get_or_404(Patient, patient_id)
    stored = save_document(request.files.get("file"))
    if not stored:
        flash(t("visits.att_bad_type"), "warning")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#documents")
    kind = request.form.get("kind")
    db.session.add(PatientAttachment(
        patient_id=patient.id, filename=stored,
        original_name=request.files["file"].filename,
        kind=kind if kind in ATTACHMENT_KINDS else "report",
        label=(request.form.get("label") or "").strip() or None,
        uploaded_by=current_user.id,
    ))
    db.session.commit()
    flash(t("visits.att_uploaded"), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#documents")


@patients_bp.route("/documents/<int:att_id>/delete", methods=["POST"])
@module_required(MODULE)
@capability_required("patient_medical")
def delete_document(att_id):
    att = db.get_or_404(PatientAttachment, att_id)
    patient_id = att.patient_id
    remove_document(att.filename)
    db.session.delete(att)
    db.session.commit()
    flash(t("visits.att_removed"), "info")
    return redirect(url_for("patients.view", patient_id=patient_id) + "#documents")


# ---------------------------------------------------------------- list -----
@patients_bp.route("/")
@module_required(MODULE)
def index():
    q = (request.args.get("q") or "").strip()
    flag = (request.args.get("flag") or "").strip()
    query = apply_patient_search(Patient.query, q)
    if flag == "teen_no_phone":
        # Reception task: active teens (≥13) with no personal phone captured yet.
        from app.models.patient import own_phone_cutoff
        query = query.filter(
            Patient.is_active.is_(True),
            Patient.date_of_birth <= own_phone_cutoff(),
            db.or_(Patient.own_phone.is_(None), Patient.own_phone == ""),
        )
    pagination = paginate(query.order_by(Patient.created_at.desc()))

    stats = {
        "total": Patient.query.count(),
        "active": Patient.query.filter_by(is_active=True).count(),
        "male": Patient.query.filter_by(gender="male").count(),
        "female": Patient.query.filter_by(gender="female").count(),
    }
    return render_template(
        "patients/list.html", patients=pagination.items, pagination=pagination,
        q=q, stats=stats, flag=flag,
    )


# ------------------------------------------------------- phone worklist ----
@patients_bp.route("/phones")
@module_required(MODULE)
def phones():
    """The reception task list: who the clinic cannot reach, and a box to fix
    it from. It replaces a notification whose only action was a count."""
    from app.utils.phonebook import worklist

    return render_template("patients/phones.html",
                           **worklist(getattr(g, "lang", "ar")))


@patients_bp.route("/phones/save", methods=["POST"])
@module_required(MODULE)
def phones_save():
    """Save one number and come straight back to the list.

    Back to the list, not into the patient's file: the whole point is working
    through thirteen of these without leaving the screen.
    """
    from app.utils.phonebook import save_number

    patient = db.get_or_404(Patient, request.form.get("patient_id", type=int))
    ok, number = save_number(
        patient, request.form.get("phone"),
        (request.form.get("target") or "own").strip(),
        request.form.get("guardian_id", type=int))
    if not ok:
        flash(t("phones.bad_number"), "danger")
    else:
        ActivityLog.record("patient.phone_add", user_id=current_user.id,
                           entity="patient", entity_id=patient.id,
                           detail=number, ip_address=client_ip())
        db.session.commit()
        flash(t("phones.saved"), "success")
    return redirect(url_for("patients.phones"))


# ------------------------------------------------------------ analytics ----
@patients_bp.route("/analytics")
@module_required(MODULE)
def analytics():
    """Patient analytics: gender / age distribution + top diagnoses in a period."""
    from datetime import datetime, timedelta

    from sqlalchemy import func

    from app.blueprints.main.routes import AGE_GROUPS, _age_group
    from app.models import Diagnosis, Visit

    days = request.args.get("days", 90, type=int)
    if days not in (30, 90, 180, 365):
        days = 90
    start = local_today() - timedelta(days=days)

    patients = Patient.query.filter_by(is_active=True).all()
    groups = {key: 0 for key, _ in AGE_GROUPS}
    male = female = 0
    for p in patients:
        if p.gender == "male":
            male += 1
        elif p.gender == "female":
            female += 1
        groups[_age_group(p.age_days)] += 1
    stats = {"total": len(patients), "male": male, "female": female, "groups": groups}

    # Top diagnoses recorded in visits within the period.
    top_dx = (
        db.session.query(Diagnosis.title, func.count().label("n"))
        .join(Visit, Diagnosis.visit_id == Visit.id)
        .filter(Visit.visit_date >= start, Diagnosis.title.isnot(None))
        .group_by(Diagnosis.title)
        .order_by(func.count().desc())
        .limit(10)
        .all()
    )
    period = {
        "days": days,
        "new_patients": Patient.query.filter(Patient.created_at >= start).count(),
        "visits": Visit.query.filter(Visit.visit_date >= start).count(),
    }
    return render_template(
        "patients/analytics.html", stats=stats,
        age_groups=[k for k, _ in AGE_GROUPS], top_dx=top_dx, period=period,
    )


# ------------------------------------------------------------ archiving ----
@patients_bp.route("/archive")
@module_required(MODULE)
def archive():
    """Archiving hub: policy, the inactivity candidate list, active/inactive
    analytics and the archived files (with restore)."""
    from app.utils.archiving import (
        archive_stats, auto_enabled, inactive_candidates, inactive_years,
    )

    years = inactive_years()
    candidates = inactive_candidates(years)
    archived = (Patient.query.filter_by(is_active=False)
                .order_by(Patient.archived_at.desc())
                .limit(200).all())
    return render_template(
        "patients/archive.html",
        stats=archive_stats(years), years=years, auto_enabled=auto_enabled(),
        candidates=candidates, archived=archived, today=local_today(),
    )


@patients_bp.route("/archive/settings", methods=["POST"])
@module_required(MODULE)
def archive_settings():
    from app.models import Setting

    years = request.form.get("years", type=int) or 3
    Setting.set("archive_inactive_years", str(min(max(years, 1), 20)))
    Setting.set("archive_auto_enabled", "1" if request.form.get("auto_enabled") else "0")
    db.session.commit()
    flash(t("archive.settings_saved"), "success")
    return redirect(url_for("patients.archive"))


@patients_bp.route("/archive/run", methods=["POST"])
@module_required(MODULE)
def archive_run():
    """Archive every current inactivity candidate now (manual sweep)."""
    from app.utils.archiving import auto_archive
    n = auto_archive()
    ActivityLog.record("patient.archive_sweep", user_id=current_user.id,
                       entity="patient", detail=str(n), ip_address=client_ip())
    db.session.commit()
    flash(t("archive.swept", n=n), "success" if n else "info")
    return redirect(url_for("patients.archive"))


@patients_bp.route("/<int:patient_id>/medications", methods=["POST"])
@module_required(MODULE)
def medication_add(patient_id):
    """Record a medicine the child is already on.

    Open to the staff who hear it, because the person who learns that a child
    takes something is usually whoever the mother is talking to. Who wrote it
    is stored, since "the mother says he takes something white" and "the
    neurologist's letter says 200mg twice a day" are not the same evidence.
    """
    from app.utils import patient_meds as meds

    patient = db.get_or_404(Patient, patient_id)
    row = meds.add(
        patient, request.form.get("name"), user=current_user,
        dose=(request.form.get("dose") or "").strip(),
        frequency=(request.form.get("frequency") or "").strip(),
        reason=(request.form.get("reason") or "").strip(),
        generic_id=request.form.get("generic_id", type=int),
        drug_id=request.form.get("drug_id", type=int),
    )
    flash(t("meds.added") if row else t("meds.need_name"),
          "success" if row else "danger")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#meds")


@patients_bp.route("/medications/<int:med_id>/stop", methods=["POST"])
@module_required(MODULE)
def medication_stop(med_id):
    """End a medicine. The row stays — see ``patient_meds.stop``."""
    from app.models import PatientMedication
    from app.utils import patient_meds as meds

    row = db.get_or_404(PatientMedication, med_id)
    meds.stop(row, user=current_user, reason=request.form.get("stop_reason"))
    flash(t("meds.stopped"), "info")
    return redirect(url_for("patients.view", patient_id=row.patient_id) + "#meds")


@patients_bp.route("/<int:patient_id>/flag", methods=["POST"])
@module_required(MODULE)
def flag_raise(patient_id):
    """Record that a family does not pay, or pays only after being chased.

    Open to the staff who meet it — reception, the doctor, the office. Taking
    it off is a different permission, so the person who wrote it while annoyed
    is not the one who lifts it.
    """
    from app.utils import patient_flags as flags

    patient = db.get_or_404(Patient, patient_id)
    level = (request.form.get("level") or "warn").strip()
    reason = (request.form.get("reason") or "").strip()
    flag = flags.raise_flag(patient.id, level, reason, user_id=current_user.id)
    if flag is None:
        flash(t("flags.need_reason"), "danger")
        return redirect(request.referrer
                        or url_for("patients.view", patient_id=patient.id))
    ActivityLog.record("patient.flag", user_id=current_user.id,
                       entity="patient", entity_id=patient.id,
                       detail=level, ip_address=client_ip())
    db.session.commit()
    flash(t("flags.raised"), "warning")
    return redirect(request.referrer
                    or url_for("patients.view", patient_id=patient.id))


@patients_bp.route("/<int:patient_id>/flag/clear", methods=["POST"])
@login_required
def flag_clear(patient_id):
    """Take the flag off when the behaviour changes — admin or finance only.

    Deliberately **not** behind ``module_required("patients")``. The
    accountant is the one person who actually knows whether the money arrived,
    and an accountant has no patients module — so gating on it meant the only
    people who could clear a payment flag were the ones who could not check
    it. Same shape as ``cashier_access``, which lets reception take money
    without handing them the whole of finance. The capability is checked
    below, so nothing is opened up: it is a narrower gate, not an absent one.
    """
    from app.utils import patient_flags as flags

    patient = db.get_or_404(Patient, patient_id)
    if not flags.can_clear(current_user):
        flash(t("flags.no_permission_clear"), "danger")
        return redirect(request.referrer
                        or url_for("patients.view", patient_id=patient.id))
    flag = flags.clear_flag(patient.id, request.form.get("clear_reason"),
                            user_id=current_user.id)
    if flag is not None:
        ActivityLog.record("patient.flag_clear", user_id=current_user.id,
                           entity="patient", entity_id=patient.id,
                           ip_address=client_ip())
        db.session.commit()
        flash(t("flags.cleared"), "success")
    return redirect(request.referrer
                    or url_for("patients.view", patient_id=patient.id))


@patients_bp.route("/<int:patient_id>/archive", methods=["POST"])
@module_required(MODULE)
def archive_one(patient_id):
    from app.utils.archiving import archive_patient
    patient = db.get_or_404(Patient, patient_id)
    if archive_patient(patient, reason="manual"):
        ActivityLog.record("patient.archive", user_id=current_user.id,
                           entity="patient", entity_id=patient.id,
                           detail=patient.patient_number, ip_address=client_ip())
        db.session.commit()
        flash(t("archive.archived_one", name=patient.display_name(g.get("lang", "ar"))),
              "info")
    return redirect(request.referrer or url_for("patients.view", patient_id=patient.id))


@patients_bp.route("/<int:patient_id>/restore", methods=["POST"])
@module_required(MODULE)
def restore_one(patient_id):
    from app.utils.archiving import restore_patient
    patient = db.get_or_404(Patient, patient_id)
    if restore_patient(patient):
        ActivityLog.record("patient.restore", user_id=current_user.id,
                           entity="patient", entity_id=patient.id,
                           detail=patient.patient_number, ip_address=client_ip())
        db.session.commit()
        flash(t("archive.restored_one", name=patient.display_name(g.get("lang", "ar"))),
              "success")
    return redirect(request.referrer or url_for("patients.view", patient_id=patient.id))


# -------------------------------------------------------------- create -----
@patients_bp.route("/new", methods=["GET", "POST"])
@module_required(MODULE)
def create():
    families = Family.query.order_by(Family.family_name).all()
    prefill_family = request.args.get("family_id", type=int)

    if request.method == "POST":
        form = _read_patient_form()
        error = _validate_patient(form, existing=None)
        if error:
            flash(error, "danger")
            return render_template(
                "patients/form.html", **_patient_chips(),
                patient=None, form=form, families=families,
                genders=GENDERS, blood_types=BLOOD_TYPES,
                suggested_number=form.get("patient_number"),
            )

        family_id = _resolve_family(form)

        patient = Patient(
            patient_number=form["patient_number"],
            reference_number=form["reference_number"] or None,
            family_id=family_id,
            full_name=form["full_name"],
            full_name_en=form["full_name_en"],
            date_of_birth=form["date_of_birth"],
            gender=form["gender"],
            national_id=form["national_id"],
            own_phone=form["own_phone"] or None,
            blood_type=form["blood_type"],
            birth_weight_kg=form["birth_weight_kg"],
            gestation_weeks=form["gestation_weeks"],
            gestation_days=form["gestation_days"],
            birth_time=form["birth_time"],
            allergies=form["allergies"],
            chronic_diseases=form["chronic_diseases"],
            notes=form["notes"],
            is_active=form["is_active"],
        )

        photo = save_patient_photo(request.files.get("photo"), _upload_dir())
        if photo:
            patient.photo = photo

        db.session.add(patient)
        db.session.flush()
        ActivityLog.record(
            "patient.create", user_id=current_user.id, entity="patient",
            entity_id=patient.id, detail=patient.patient_number,
            ip_address=client_ip(),
        )
        db.session.commit()
        flash(t("patients.created"), "success")
        return redirect(url_for("patients.view", patient_id=patient.id))

    return render_template(
        "patients/form.html", **_patient_chips(),
        patient=None, form={"is_active": True}, families=families,
        genders=GENDERS, blood_types=BLOOD_TYPES,
        suggested_number=generate_patient_number(),
        prefill_family=prefill_family,
    )


# ---------------------------------------------------------------- view -----
@patients_bp.route("/<int:patient_id>")
@module_required(MODULE)
def view(patient_id):
    from app.models import Invoice, PayerEntity, Prescription
    from app.utils import ai as ai_utils
    from app.utils import lab_series, series

    patient = db.get_or_404(Patient, patient_id)
    ai_patient = (current_user.can_access("ai") and ai_utils.is_ready()
                  and ai_utils.patient_context_enabled())
    # The discussion card needs both: the record still leaves the building, and
    # a differential is its own decision on top of that.
    ai_discuss = ai_patient and ai_utils.discussion_enabled()

    prescriptions = (Prescription.query.filter_by(patient_id=patient.id)
                     .order_by(Prescription.rx_date.desc(), Prescription.id.desc()).all())
    invoices = (Invoice.query.filter_by(patient_id=patient.id)
                .order_by(Invoice.invoice_date.desc(), Invoice.id.desc()).all())
    fin = {
        "total": round(sum(i.total for i in invoices), 2),
        "paid": round(sum(i.paid for i in invoices), 2),
        "balance": round(sum(i.balance for i in invoices), 2),
    }
    from app.utils.consent import all_statements
    # Device studies — echo, audiometry, ECG, spirometry. They were in three
    # places on this screen: a table on the overview, a second list at the
    # bottom of the visits tab, and nowhere they could be read per device.
    # Three copies of one list is three chances to disagree.
    from app.utils.studies import patient_studies

    # Case history brought across from the program the clinic used before. The
    # vaccinations among it became real vaccination records, so they show on
    # that tab; the plain services — كشف, إستشارة, most of a real export — had
    # nowhere to appear at all, which made "the clinic does not lose its
    # history" only half true.
    from app.models import ImportedService
    imported = (ImportedService.query
                .filter_by(patient_id=patient.id)
                .order_by(ImportedService.service_date.desc(),
                          ImportedService.id.desc()).all())

    # Brothers and sisters who are already patients here. "Add sibling" only
    # ever led to a blank new-patient form, which in a clinic that has been
    # open a while is the wrong half of the problem: the sibling usually
    # exists already, and creating him again makes two files for one child.
    from app.utils.siblings import suggest_siblings

    from app.utils import patient_flags as flags
    from app.utils import patient_meds as meds

    # The vaccination tab asks a different question from the certificate.
    # The certificate is a record to hand a family — what the child *had*.
    # This is a glance while somebody is looking at the file: where does this
    # child stand, and what is owed next. So it is built from the plan rather
    # than from the raw dose rows, which could say what happened and never
    # what is coming.
    from app.utils.course_state import annotate as annotate_courses
    from app.utils.vaccines import (certificate_cards, next_due_dose,
                                    patient_plan)

    vlang = getattr(g, "lang", "ar")
    vaccine_plan = annotate_courses(patient_plan(patient, vlang))
    vaccine_next = next_due_dose(vaccine_plan)

    return render_template(
        "patients/profile.html",
        vaccine_cards=certificate_cards(vaccine_plan),
        vaccine_plan=vaccine_plan,
        vaccine_next=vaccine_next,
        studies=patient_studies(patient, getattr(g, "lang", "ar")),
        imported=imported,
        payment_flag=flags.active(patient.id),
        medications=meds.history(patient),
        payment_flag_history=flags.history(patient.id),
        can_clear_flag=flags.can_clear(current_user),
        sibling_hints=suggest_siblings(patient),
        patient=patient,
        relations=PARENT_RELATIONS,
        # The child's mouth, when the clinic does dentistry. A summary and a
        # way in — the chart itself is a screen of its own and belongs there,
        # not copied into this file. Everything about dentistry was built and
        # nothing linked to it: the chart, the plans and the per-tooth history
        # all existed, and no screen in the program pointed at any of them.
        dental=_dental_summary(patient),
        consent_types=CONSENT_TYPES,
        consent_statements=all_statements(),
        categories=CLIENT_CATEGORIES,
        payers=PayerEntity.query.filter_by(is_active=True).order_by(PayerEntity.name).all(),
        ai_patient=ai_patient, ai_discuss=ai_discuss,
        prescriptions=prescriptions, invoices=invoices, fin=fin,
        growth_alert=_growth_concern(patient),
        # One reading across every visit — labs, device studies and specialty
        # panels in one list, and the EF from an echo on the same line as the
        # EF the cardiologist wrote. See app/utils/series.py.
        lab_series=series.curves_for(patient.id, getattr(g, "lang", "ar")),
        lab_latest=lab_series.latest_values(patient.id, getattr(g, "lang", "ar")),
    )


@patients_bp.route("/<int:patient_id>/report")
@module_required(MODULE)
def report(patient_id):
    """Generate a comprehensive medical report for the case: demographics,
    problem list, growth, vaccination status, recent visits, current meds — all
    assembled from the record. The doctor can edit any section inline in the
    browser and print it (editing is browser-side; the record isn't changed)."""
    from app.models import GrowthRecord, Prescription, Visit
    from app.utils.vaccines import patient_plan, plan_summary

    patient = db.get_or_404(Patient, patient_id)

    problems = [p for p in getattr(patient, "problems", [])
                if p.status == "active"]
    visits = (Visit.query.filter_by(patient_id=patient.id)
              .order_by(Visit.visit_date.desc(), Visit.id.desc()).limit(10).all())
    latest_growth = (GrowthRecord.query.filter_by(patient_id=patient.id)
                     .order_by(GrowthRecord.record_date.desc(),
                               GrowthRecord.id.desc()).first())
    try:
        vac = plan_summary(patient_plan(patient, getattr(g, "lang", "ar")))
    except Exception:  # noqa: BLE001
        vac = None
    latest_rx = (Prescription.query.filter_by(patient_id=patient.id)
                 .order_by(Prescription.rx_date.desc(), Prescription.id.desc()).first())

    return render_template(
        "patients/report.html", patient=patient, problems=problems,
        visits=visits, latest_growth=latest_growth, vac=vac,
        latest_rx=latest_rx, growth_alert=_growth_concern(patient),
        generated_by=current_user, today=local_today(),
    )


def _growth_concern(patient):
    """If the patient's latest growth measurement falls in a caution/alert band
    (|z|>2), return a compact dict so the profile can flag it prominently."""
    from app.models import GrowthRecord
    from app.utils.growth import (
        INDICATORS, age_for, compute_point, reference_for, status_for_z,
    )

    rec = (GrowthRecord.query.filter_by(patient_id=patient.id)
           .order_by(GrowthRecord.record_date.desc(), GrowthRecord.id.desc()).first())
    if rec is None:
        return None
    ref = reference_for(patient)
    # A premature child scored at their birthday age reads as small when they
    # are on course, and "small" on a growth chart is what starts a workup.
    age = age_for(patient, rec.record_date)
    worst = None
    for ind, meta in INDICATORS.items():
        value = getattr(rec, meta["field"], None)
        if not value:
            continue
        pt = compute_point(ref, ind, patient.gender,
                           patient.date_of_birth, rec.record_date, value,
                           age_months=age["months"])
        if not pt or pt.get("z") is None:
            continue
        status = status_for_z(pt["z"])
        if status in ("caution", "alert"):
            cand = {"indicator": ind, "z": pt["z"], "percentile": pt["percentile"],
                    "status": status, "date": rec.record_date.isoformat(),
                    "corrected": age["corrected"]}
            if worst is None or abs(pt["z"]) > abs(worst["z"]):
                worst = cand
    return worst


# ---------------------------------------------------------------- edit -----
@patients_bp.route("/<int:patient_id>/edit", methods=["GET", "POST"])
@module_required(MODULE)
def edit(patient_id):
    patient = db.get_or_404(Patient, patient_id)
    families = Family.query.order_by(Family.family_name).all()

    if request.method == "POST":
        form = _read_patient_form()
        error = _validate_patient(form, existing=patient)
        if error:
            flash(error, "danger")
            return render_template(
                "patients/form.html", **_patient_chips(),
                patient=patient, form=form, families=families,
                genders=GENDERS, blood_types=BLOOD_TYPES,
                suggested_number=form.get("patient_number"),
            )

        patient.patient_number = form["patient_number"]
        patient.reference_number = form["reference_number"] or None
        patient.family_id = _resolve_family(form)
        patient.full_name = form["full_name"]
        patient.full_name_en = form["full_name_en"]
        patient.date_of_birth = form["date_of_birth"]
        patient.gender = form["gender"]
        patient.national_id = form["national_id"]
        patient.own_phone = form["own_phone"] or None
        patient.blood_type = form["blood_type"]
        patient.birth_weight_kg = form["birth_weight_kg"]
        patient.gestation_weeks = form["gestation_weeks"]
        patient.gestation_days = form["gestation_days"]
        patient.birth_time = form["birth_time"]
        patient.allergies = form["allergies"]
        patient.chronic_diseases = form["chronic_diseases"]
        patient.notes = form["notes"]
        patient.is_active = form["is_active"]

        new_photo = save_patient_photo(request.files.get("photo"), _upload_dir())
        if new_photo:
            delete_patient_photo(patient.photo, _upload_dir())
            patient.photo = new_photo

        ActivityLog.record(
            "patient.update", user_id=current_user.id, entity="patient",
            entity_id=patient.id, detail=patient.patient_number,
            ip_address=client_ip(),
        )
        db.session.commit()
        flash(t("patients.updated"), "success")
        return redirect(url_for("patients.view", patient_id=patient.id))

    form = {
        "patient_number": patient.patient_number,
        "reference_number": patient.reference_number or "",
        "family_id": patient.family_id,
        "full_name": patient.full_name,
        "full_name_en": patient.full_name_en or "",
        "date_of_birth": patient.date_of_birth.isoformat() if patient.date_of_birth else "",
        "gender": patient.gender,
        "national_id": patient.national_id or "",
        "own_phone": patient.own_phone or "",
        "blood_type": patient.blood_type or "",
        "birth_weight_kg": patient.birth_weight_kg if patient.birth_weight_kg is not None else "",
        "gestation_weeks": patient.gestation_weeks if patient.gestation_weeks is not None else "",
        "gestation_days": patient.gestation_days if patient.gestation_days is not None else "",
        # `HH:MM` — what an <input type="time"> round-trips. Blank when
        # nobody recorded it, so re-saving an edit form does not invent one.
        "birth_time": (patient.birth_time.strftime("%H:%M")
                       if patient.birth_time else ""),
        "allergies": patient.allergies or "",
        "chronic_diseases": patient.chronic_diseases or "",
        "notes": patient.notes or "",
        "is_active": patient.is_active,
    }
    return render_template(
        "patients/form.html", **_patient_chips(),
        patient=patient, form=form, families=families,
        genders=GENDERS, blood_types=BLOOD_TYPES,
        suggested_number=patient.patient_number,
    )


# -------------------------------------------------------------- delete -----
@patients_bp.route("/<int:patient_id>/delete", methods=["POST"])
@module_required(MODULE)
def delete(patient_id):
    patient = db.get_or_404(Patient, patient_id)
    number = patient.patient_number
    delete_patient_photo(patient.photo, _upload_dir())
    db.session.delete(patient)
    ActivityLog.record(
        "patient.delete", user_id=current_user.id, entity="patient",
        entity_id=patient_id, detail=number, ip_address=client_ip(),
    )
    db.session.commit()
    flash(t("patients.deleted"), "info")
    return redirect(url_for("patients.index"))


# ----------------------------------------------- membership / coverage -----
@patients_bp.route("/<int:patient_id>/coverage/new", methods=["POST"])
@module_required(MODULE)
def add_coverage(patient_id):
    from datetime import datetime as _dt

    from app.models import PatientCoverage, PayerEntity

    patient = db.get_or_404(Patient, patient_id)
    payer = db.session.get(PayerEntity, request.form.get("payer_id", type=int))
    if payer is None:
        flash(t("coverage.need_payer"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#coverage")

    def _exp():
        raw = (request.form.get("expiry_date") or "").strip()
        try:
            return _dt.strptime(raw, "%Y-%m-%d").date() if raw else None
        except ValueError:
            return None

    card = (request.form.get("membership_number") or "").strip() or None
    expiry = _exp()
    targets = [patient]
    # Optionally apply the same entity to siblings (each keeps its own card).
    if request.form.get("apply_siblings"):
        targets += patient.siblings

    for tgt in targets:
        # Per-patient card number only for the main patient; siblings share the
        # entity but get their own (blank) card to fill later.
        db.session.add(PatientCoverage(
            patient_id=tgt.id, payer_id=payer.id,
            membership_number=card if tgt.id == patient.id else None,
            expiry_date=expiry, is_active=True,
        ))
    db.session.commit()
    flash(t("coverage.added"), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#coverage")


@patients_bp.route("/coverage/<int:coverage_id>/delete", methods=["POST"])
@module_required(MODULE)
def delete_coverage(coverage_id):
    from app.models import PatientCoverage

    cov = db.get_or_404(PatientCoverage, coverage_id)
    pid = cov.patient_id
    db.session.delete(cov)
    db.session.commit()
    flash(t("coverage.removed"), "info")
    return redirect(url_for("patients.view", patient_id=pid) + "#coverage")


# ------------------------------------------------------ parents (family) ---
@patients_bp.route("/<int:patient_id>/parents/new", methods=["POST"])
@module_required(MODULE)
def add_parent(patient_id):
    patient = db.get_or_404(Patient, patient_id)

    # Ensure the patient has a family to attach the parent to.
    if patient.family_id is None:
        family = Family(family_name=patient.full_name)
        db.session.add(family)
        db.session.flush()
        patient.family_id = family.id

    relation = (request.form.get("relation") or "father").strip()
    full_name = (request.form.get("full_name") or "").strip()
    category = (request.form.get("client_category") or "normal").strip()

    if not full_name:
        flash(t("common.required") + ": " + t("patients.parent_name"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id))

    parent = Parent(
        family_id=patient.family_id,
        relation=relation if Parent.valid_relation(relation) else "father",
        full_name=full_name,
        full_name_en=(request.form.get("full_name_en") or "").strip(),
        national_id=(request.form.get("national_id") or "").strip(),
        phone=(request.form.get("phone") or "").strip(),
        phone_alt=(request.form.get("phone_alt") or "").strip(),
        email=(request.form.get("email") or "").strip(),
        occupation=(request.form.get("occupation") or "").strip(),
        nationality=(request.form.get("nationality") or "").strip(),
        address=(request.form.get("address") or "").strip(),
        client_category=category if Parent.valid_category(category) else "normal",
        is_primary_contact=bool(request.form.get("is_primary_contact")),
    )
    db.session.add(parent)
    ActivityLog.record(
        "parent.create", user_id=current_user.id, entity="parent",
        entity_id=patient.family_id, detail=full_name, ip_address=client_ip(),
    )
    db.session.commit()
    flash(t("patients.parent_added"), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#family")


@patients_bp.route("/parents/<int:parent_id>/edit", methods=["POST"])
@module_required(MODULE)
def edit_parent(parent_id):
    parent = db.get_or_404(Parent, parent_id)
    patient_id = request.form.get("patient_id", type=int)

    full_name = (request.form.get("full_name") or "").strip()
    if not full_name:
        flash(t("common.required") + ": " + t("patients.parent_name"), "danger")
        return redirect(url_for("patients.view", patient_id=patient_id) + "#family")

    relation = (request.form.get("relation") or parent.relation).strip()
    category = (request.form.get("client_category") or parent.client_category).strip()
    parent.relation = relation if Parent.valid_relation(relation) else parent.relation
    parent.full_name = full_name
    parent.full_name_en = (request.form.get("full_name_en") or "").strip()
    parent.national_id = (request.form.get("national_id") or "").strip()
    parent.phone = (request.form.get("phone") or "").strip()
    parent.phone_alt = (request.form.get("phone_alt") or "").strip()
    parent.email = (request.form.get("email") or "").strip()
    parent.occupation = (request.form.get("occupation") or "").strip()
    parent.nationality = (request.form.get("nationality") or "").strip()
    parent.address = (request.form.get("address") or "").strip()
    parent.client_category = category if Parent.valid_category(category) else parent.client_category
    parent.is_primary_contact = bool(request.form.get("is_primary_contact"))

    ActivityLog.record(
        "parent.update", user_id=current_user.id, entity="parent",
        entity_id=parent.id, detail=full_name, ip_address=client_ip(),
    )
    db.session.commit()
    flash(t("patients.parent_updated"), "success")
    return redirect(url_for("patients.view", patient_id=patient_id) + "#family")


@patients_bp.route("/parents/<int:parent_id>/delete", methods=["POST"])
@module_required(MODULE)
def delete_parent(parent_id):
    parent = db.get_or_404(Parent, parent_id)
    patient_id = request.form.get("patient_id", type=int)
    db.session.delete(parent)
    db.session.commit()
    flash(t("patients.parent_removed"), "info")
    if patient_id:
        return redirect(url_for("patients.view", patient_id=patient_id) + "#family")
    return redirect(url_for("patients.index"))


# ------------------------------------------------------ problem list -------
def _parse_date(name):
    raw = (request.form.get(name) or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


def _icd_version(raw):
    """One of the classifications this program knows, or None."""
    from app.utils.icd import VERSIONS

    value = (raw or "").strip()
    return value if value in VERSIONS else None


@patients_bp.route("/<int:patient_id>/problems", methods=["POST"])
@module_required(MODULE)
def add_problem(patient_id):
    patient = db.get_or_404(Patient, patient_id)
    title = (request.form.get("title") or "").strip()
    if not title:
        flash(t("common.required") + ": " + t("problems.title"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#problems")
    db.session.add(PatientProblem(
        patient_id=patient.id, title=title,
        title_en=(request.form.get("title_en") or "").strip() or None,
        icd_code=(request.form.get("icd_code") or "").strip() or None,
        # Checked against the versions the program knows rather than trusted:
        # the picker fills it, but the code box beside it can be typed by
        # hand, and a version nobody recognises is worse than none.
        icd_version=_icd_version(request.form.get("icd_version")),
        onset_date=_parse_date("onset_date"),
        notes=(request.form.get("notes") or "").strip() or None,
    ))
    ActivityLog.record("problem.add", user_id=current_user.id, entity="patient",
                       entity_id=patient.id, detail=title, ip_address=client_ip())
    db.session.commit()
    flash(t("problems.added"), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#problems")


@patients_bp.route("/problems/<int:problem_id>/toggle", methods=["POST"])
@module_required(MODULE)
def toggle_problem(problem_id):
    prob = db.get_or_404(PatientProblem, problem_id)
    if prob.status == "active":
        prob.status = "resolved"
        prob.resolved_date = local_today()
    else:
        prob.status = "active"
        prob.resolved_date = None
    db.session.commit()
    flash(t("problems.updated"), "success")
    return redirect(url_for("patients.view", patient_id=prob.patient_id) + "#problems")


@patients_bp.route("/problems/<int:problem_id>/delete", methods=["POST"])
@module_required(MODULE)
def delete_problem(problem_id):
    prob = db.get_or_404(PatientProblem, problem_id)
    pid = prob.patient_id
    db.session.delete(prob)
    db.session.commit()
    flash(t("problems.deleted"), "info")
    return redirect(url_for("patients.view", patient_id=pid) + "#problems")


# --------------------------------------------------------- consent ---------
@patients_bp.route("/<int:patient_id>/consents", methods=["POST"])
@module_required(MODULE)
def add_consent(patient_id):
    patient = db.get_or_404(Patient, patient_id)
    guardian = (request.form.get("guardian_name") or "").strip()
    ctype = (request.form.get("consent_type") or "general").strip()
    if not guardian:
        flash(t("common.required") + ": " + t("consent.guardian_name"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#consent")
    # Through the one writer, not a second hand-built row: this screen and the
    # visit room were both creating consents, and only one of them would have
    # got the per-kind wording — leaving whichever was used that day to decide
    # what a guardian appears to have agreed to.
    from app.utils.consent import record as record_consent

    record_consent(
        patient, ctype, guardian,
        relation=(request.form.get("guardian_relation") or "").strip() or None,
        id_no=(request.form.get("guardian_id_no") or "").strip() or None,
        statement=(request.form.get("statement") or "").strip() or None,
        notes=(request.form.get("notes") or "").strip() or None,
        user_id=current_user.id,
        on_date=_parse_date("signed_date") or local_today(),
    )
    ActivityLog.record("consent.add", user_id=current_user.id, entity="patient",
                       entity_id=patient.id, detail=ctype, ip_address=client_ip())
    db.session.commit()
    flash(t("consent.added"), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#consent")


@patients_bp.route("/consents/<int:consent_id>/delete", methods=["POST"])
@module_required(MODULE)
def delete_consent(consent_id):
    """Remove a consent entirely. Admin only, and it is not the ordinary path.

    Withdrawing is what a guardian does and what :func:`withdraw_consent`
    records. This is for the other thing: a consent written on the wrong
    child, which is not a fact about anybody and must not sit in their file
    being true. Anyone who can edit a patient could do this before, which
    meant the difference between the two was a button label.
    """
    if not current_user.is_admin:
        abort(403)
    c = db.get_or_404(Consent, consent_id)
    pid = c.patient_id
    ActivityLog.record("consent.delete", user_id=current_user.id,
                       entity="patient", entity_id=pid,
                       ip_address=client_ip())
    db.session.delete(c)
    db.session.commit()
    flash(t("consent.deleted"), "info")
    return redirect(url_for("patients.view", patient_id=pid) + "#consent")


@patients_bp.route("/consents/<int:consent_id>/withdraw", methods=["POST"])
@module_required(MODULE)
def withdraw_consent(consent_id):
    """The guardian has withdrawn it. The row stays; it is marked and dated.

    The statement they signed says they may — *"ولي أن أسحب موافقتي في أي
    وقت"* — and the program could only delete, which is a different thing. The
    consent was given; that remains true. What changed is that it no longer
    stands, and when it stopped standing is exactly the fact somebody will
    need later.

    Reversible for the same reason a referral is: a withdrawal recorded
    against the wrong consent, at a desk, with a queue waiting.
    """
    from app.utils.clock import local_now

    c = db.get_or_404(Consent, consent_id)
    if c.is_withdrawn:
        c.withdrawn_at = None
        c.withdrawn_reason = None
        c.withdrawn_by = None
        action, message = "consent.withdraw_undo", "consent.withdraw_undone"
    else:
        c.withdrawn_at = local_now().replace(tzinfo=None)
        c.withdrawn_reason = (request.form.get("reason") or "").strip() or None
        c.withdrawn_by = current_user.id
        action, message = "consent.withdraw", "consent.withdrawn"
    ActivityLog.record(action, user_id=current_user.id, entity="patient",
                       entity_id=c.patient_id, ip_address=client_ip())
    db.session.commit()
    flash(t(message), "info")
    return redirect(
        url_for("patients.view", patient_id=c.patient_id) + "#consent")


@patients_bp.route("/consents/<int:consent_id>/signature", methods=["POST"])
@module_required(MODULE)
def consent_signature(consent_id):
    """Attach the evidence: the scanned paper, or a signature drawn on screen.

    Both arrive here because they are the same fact recorded two ways, and a
    file must be able to say which one it holds — see `Consent.signature_kind`.

    **Neither path trusts what it was sent.** The upload goes through
    `save_document`, which decides the type from the bytes and not the name.
    The drawn one is a base64 image from a canvas, and it is decoded and then
    put through the same sniffing: a data URL claiming to be a PNG is a string
    somebody typed, and this writes files into a folder the browser serves.
    """
    import base64
    import binascii
    import os
    import uuid

    from app.utils.clock import local_now
    from app.utils.uploads import (ALLOWED_DOC_EXTENSIONS, docs_dir,
                                   save_document, sniff_ext)

    c = db.get_or_404(Consent, consent_id)
    back = url_for("patients.view", patient_id=c.patient_id) + "#consent"

    drawn = (request.form.get("drawn") or "").strip()
    if drawn:
        # `data:image/png;base64,….` — the header is discarded rather than
        # believed; what the bytes are is decided below.
        payload = drawn.split(",", 1)[-1]
        try:
            raw = base64.b64decode(payload, validate=True)
        except (binascii.Error, ValueError):
            flash(t("visits.att_bad_type"), "warning")
            return redirect(back)
        ext = sniff_ext(raw[:64])
        if ext not in ALLOWED_DOC_EXTENSIONS:
            flash(t("visits.att_bad_type"), "warning")
            return redirect(back)
        stored = f"{uuid.uuid4().hex}.{ext}"
        os.makedirs(docs_dir(), exist_ok=True)
        with open(os.path.join(docs_dir(), stored), "wb") as fh:
            fh.write(raw)
        kind = "drawn"
    else:
        stored = save_document(request.files.get("file"))
        if not stored:
            flash(t("visits.att_bad_type"), "warning")
            return redirect(back)
        kind = "paper"

    c.signature_file = stored
    c.signature_kind = kind
    c.signature_at = local_now().replace(tzinfo=None)
    ActivityLog.record(f"consent.signature.{kind}", user_id=current_user.id,
                       entity="patient", entity_id=c.patient_id,
                       ip_address=client_ip())
    db.session.commit()
    flash(t("consent.signature_saved"), "success")
    return redirect(back)


@patients_bp.route("/consents/<int:consent_id>/print")
@module_required(MODULE)
def print_consent(consent_id):
    from app.utils.consent import statement_for

    c = db.get_or_404(Consent, consent_id)
    return render_template("patients/consent_print.html", c=c, patient=c.patient,
                           statement_for=statement_for)


# ----------------------------------------------------------- families ------
@patients_bp.route("/families/search")
@module_required(MODULE)
def family_search():
    """Autocomplete for picking a family by name or number (booking/new patient)."""
    q = (request.args.get("q") or "").strip()
    if len(q) < 1:
        return jsonify({"families": []})
    like = f"%{q}%"
    rows = (
        Family.query.filter(db.or_(
            Family.family_name.ilike(like),
            Family.family_name_en.ilike(like),
            Family.family_number.ilike(like),
        )).order_by(Family.family_name).limit(15).all()
    )
    return jsonify({"families": [
        {"id": f.id, "name": f.display_name(),
         "number": f.family_number or "",
         "count": len(f.patients)} for f in rows
    ]})


@patients_bp.route("/families/<int:family_id>/edit", methods=["POST"])
@module_required(MODULE)
def family_update(family_id):
    """Rename / edit a family's details."""
    family = db.get_or_404(Family, family_id)
    name = (request.form.get("family_name") or "").strip()
    if not name:
        flash(t("common.required") + ": " + t("patients.family_name"), "danger")
    else:
        family.family_name = name
        family.family_name_en = (request.form.get("family_name_en") or "").strip() or None
        family.notes = (request.form.get("notes") or "").strip() or None
        db.session.commit()
        flash(t("patients.family_updated"), "success")
    patient_id = request.form.get("patient_id", type=int)
    if patient_id:
        return redirect(url_for("patients.view", patient_id=patient_id) + "#family")
    return redirect(url_for("patients.index"))


# --------------------------------------------------------------- helpers ---
def _patient_chips():
    """The one-click allergy and chronic-illness lists, for the form.

    Handed to every render of it rather than to one: the create screen and the
    edit screen are the same template, and a helper that reached only the
    first is how a feature ends up existing on half the paths that show it.
    """
    from app.utils import patient_chips

    return {"allergy_chips": patient_chips.chips("allergy"),
            "chronic_chips": patient_chips.chips("chronic")}


def _read_patient_form():
    return {
        "patient_number": (request.form.get("patient_number") or "").strip(),
        "auto_number": bool(request.form.get("auto_number")),
        "reference_number": (request.form.get("reference_number") or "").strip(),
        "full_name": (request.form.get("full_name") or "").strip(),
        "full_name_en": (request.form.get("full_name_en") or "").strip(),
        "date_of_birth_raw": (request.form.get("date_of_birth") or "").strip(),
        "gender": (request.form.get("gender") or "").strip(),
        "national_id": (request.form.get("national_id") or "").strip(),
        "own_phone": (request.form.get("own_phone") or "").strip(),
        "blood_type": (request.form.get("blood_type") or "").strip(),
        # What the child arrived with. Blank stays blank — an unrecorded birth
        # weight is not a zero, and an unrecorded gestation is not "term".
        "birth_weight_kg": (request.form.get("birth_weight_kg") or "").strip(),
        "gestation_weeks": (request.form.get("gestation_weeks") or "").strip(),
        "gestation_days": (request.form.get("gestation_days") or "").strip(),
        "birth_time": (request.form.get("birth_time") or "").strip(),
        "allergies": (request.form.get("allergies") or "").strip(),
        "chronic_diseases": (request.form.get("chronic_diseases") or "").strip(),
        "notes": (request.form.get("notes") or "").strip(),
        "is_active": bool(request.form.get("is_active")),
        # Family selection.
        "family_id": request.form.get("family_id", type=int),
        "new_family_name": (request.form.get("new_family_name") or "").strip(),
    }


def _dental_summary(patient):
    """What this child's mouth needs, in the few lines a file tab can hold.

    ``None`` when the clinic does not do dentistry, so the tab is absent
    rather than empty — an empty tab labelled "teeth" on every paediatric file
    is furniture, and the module is off by default precisely because a
    paediatric clinic is not a dental one.
    """
    from app.utils.facility import module_enabled

    if not module_enabled("dentistry"):
        return None

    from app.models import TreatmentPlan
    from app.models.dental import chart_for, outstanding

    drawn = chart_for(patient.id)
    plans = (TreatmentPlan.query
             .filter(TreatmentPlan.patient_id == patient.id,
                     TreatmentPlan.status.in_(("draft", "accepted")))
             .order_by(TreatmentPlan.id.desc()).all())
    return {"needs": outstanding(drawn), "plans": plans,
            "recorded": sum(1 for slots in drawn.values() if slots)}


def _validate_patient(form, existing):
    """Validate, parse the date, and assign the file number. Returns error str."""
    if not form["full_name"]:
        return t("common.required") + ": " + t("patients.full_name")
    if not Patient.valid_gender(form["gender"]):
        return t("common.required") + ": " + t("patients.gender")

    # Date of birth.
    if not form["date_of_birth_raw"]:
        return t("common.required") + ": " + t("patients.dob")
    try:
        form["date_of_birth"] = datetime.strptime(
            form["date_of_birth_raw"], "%Y-%m-%d"
        ).date()
    except ValueError:
        return t("patients.invalid_date")

    # Birth weight and gestation. Both optional; both refused rather than
    # silently coerced, because a birth weight of 32 is a gestation typed into
    # the wrong box and a chart would plot it without complaint.
    error = _read_birth_facts(form)
    if error:
        return error

    # File number: auto-generate or validate uniqueness of the manual one.
    if form["auto_number"] or not form["patient_number"]:
        if existing is not None:
            form["patient_number"] = existing.patient_number
        else:
            form["patient_number"] = generate_patient_number()
    else:
        dup = Patient.query.filter_by(patient_number=form["patient_number"])
        if existing is not None:
            dup = dup.filter(Patient.id != existing.id)
        if dup.first() is not None:
            return t("patients.number_taken")

    return None


def _read_birth_facts(form, ranges=((0.3, 7.0), (22, 45))):
    """Parse birth weight and gestation into the form, or return an error.

    The bounds are deliberately wide and deliberately present. Wide, because a
    400-gram survivor and a 6-kilo baby are both real and a program that
    refused them would be wrong about a child who exists. Present, because the
    two boxes sit next to each other: "36" typed into the weight box is a
    gestation, and a growth chart would plot it as a six-year-old's weight at
    birth without a murmur.
    """
    (w_lo, w_hi), (g_lo, g_hi) = ranges

    form["birth_weight_kg"] = None if not form.get("birth_weight_kg") else \
        _float_or(form["birth_weight_kg"])
    if form["birth_weight_kg"] is not None and \
            not (w_lo <= form["birth_weight_kg"] <= w_hi):
        return t("patients.birth_weight_range")

    form["gestation_weeks"] = None if not form.get("gestation_weeks") else \
        _int_or(form["gestation_weeks"])
    form["gestation_days"] = None if not form.get("gestation_days") else \
        _int_or(form["gestation_days"])

    if form["gestation_weeks"] is not None and \
            not (g_lo <= form["gestation_weeks"] <= g_hi):
        return t("patients.gestation_range")
    if form["gestation_days"] is not None and not (0 <= form["gestation_days"] <= 6):
        return t("patients.gestation_days_range")
    # Days without weeks is half a gestation and means nothing on its own.
    if form["gestation_weeks"] is None:
        form["gestation_days"] = None
    elif form["gestation_days"] is None:
        form["gestation_days"] = 0

    # The hour of birth. Blank becomes ``None`` and never midnight: the whole
    # point of recording it is that an assumed hour is wrong by up to twelve,
    # and a stored 00:00 is indistinguishable from a baby actually born then.
    raw_time = form.get("birth_time") or ""
    if not raw_time:
        form["birth_time"] = None
    else:
        try:
            hours, minutes = raw_time.split(":")[:2]
            form["birth_time"] = dtime(int(hours), int(minutes))
        except (ValueError, TypeError):
            return t("patients.birth_time_bad")
    return None


def _float_or(raw, default=None):
    try:
        return float(str(raw).strip())
    except (TypeError, ValueError):
        return default


def _int_or(raw, default=None):
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError):
        return default


def _resolve_family(form):
    """Return a family id based on the form: existing, new, or None.

    A typed name is matched against the register before anything is created.
    A clinic found its imported siblings in two families and did the obvious
    thing — opened each child and typed the family name they wanted on both —
    which produced two family records with one name on them and left the
    children as far apart as before. Typing is how somebody says what the
    family is *called*; it should not be how a duplicate gets made.

    Matched on the folded name, like every other Arabic name comparison here:
    "أحمد" and "احمد" are one household, and so are "Mohammed Khafaga" and
    "mohammed khafaga". Folding is about spelling and nothing else — two
    families that wrote different names stay two families.
    """
    typed = form.get("new_family_name")
    if typed:
        from app.utils.history_import import normalise_arabic

        key = normalise_arabic(typed)
        for existing in Family.query.all():
            if normalise_arabic(existing.family_name or "") == key:
                return existing.id
        family = Family(family_name=typed)
        db.session.add(family)
        db.session.flush()
        return family.id
    return form.get("family_id") or None


# ------------------------------------------------------- bulk import -------
MAX_PREVIEW_ROWS = 200

# The one value in the doctor dropdown that is not a user id. Deliberately not
# a number, so it can never be confused with one, and deliberately checked
# against a constant in both places rather than spelt out twice.
CREATE_DOCTOR = "new"


def _import_tmp_dir():
    path = os.path.join(current_app.instance_path, "import_tmp")
    os.makedirs(path, exist_ok=True)
    return path


def _analyze_rows(rows):
    """Validate parsed rows without writing anything (for the preview step).

    Returns (preview, valid_count) where preview is a per-row list with the
    resolved values and an ``ok``/``error`` status + reason.
    """
    from app.utils import patient_dedupe

    known = patient_dedupe.index()
    in_file = {}
    preview = []
    valid = 0
    for offset, row in enumerate(rows):
        line = offset + 2  # +1 header, 1-based
        name = (str(row.get("full_name")).strip() if row.get("full_name") else "")
        gender = parse_gender(row.get("gender"))
        dob = parse_date(row.get("date_of_birth"))

        reason = None
        if not name:
            reason = t("import.err_name")
        elif gender is None:
            reason = t("import.err_gender")
        elif dob is None:
            reason = t("import.err_dob")

        # Already in the register, or listed twice in this sheet. Not an
        # error — a second upload is the normal way a clinic adds a few
        # months — but it has to be visible *before* the import runs, or the
        # numbers afterwards are a surprise.
        duplicate = False
        if reason is None:
            duplicate = (patient_dedupe.match(row, dob, known) is not None
                         or patient_dedupe.match(row, dob, in_file) is not None)
            if not duplicate:
                for key in patient_dedupe.row_keys(row, dob):
                    in_file.setdefault(key, line)
                valid += 1
        preview.append({
            "line": line,
            "name": name or "—",
            "name_en": (row.get("full_name_en") or "").strip() or None,
            "gender": gender,
            "dob": dob.isoformat() if dob else (row.get("date_of_birth") or "—"),
            "family": (row.get("family_name") or "").strip() or None,
            "parent": (row.get("parent_name") or "").strip() or None,
            "ok": reason is None and not duplicate,
            "duplicate": duplicate,
            "reason": reason or (t("import.already_here") if duplicate else None),
        })
    return preview, valid


@patients_bp.route("/import", methods=["GET", "POST"])
@module_required(MODULE)
def bulk_import():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash(t("import.no_file"), "danger")
            return redirect(url_for("patients.bulk_import"))
        if not allowed_import_file(file.filename):
            flash(t("import.bad_format"), "danger")
            return redirect(url_for("patients.bulk_import"))

        headers, data_rows, error = read_matrix(file)
        if error == "unreadable":
            flash(t("import.unreadable"), "danger")
            return redirect(url_for("patients.bulk_import"))
        if error or not data_rows:
            flash(t("import.empty"), "warning")
            return redirect(url_for("patients.bulk_import"))

        # Stash the raw sheet (headers + rows) so the mapping step can rebuild
        # records however the user maps the columns. Dates -> ISO for JSON.
        token = uuid.uuid4().hex
        with open(os.path.join(_import_tmp_dir(), f"{token}.json"), "w",
                  encoding="utf-8") as fh:
            json.dump({"headers": headers, "rows": data_rows, "filename": file.filename},
                      fh, ensure_ascii=False, default=_json_cell)

        return render_template(
            "patients/import_map.html",
            token=token, headers=headers, filename=file.filename,
            fields=import_fields(), required_keys=REQUIRED_KEYS,
            guess=guess_mapping(headers),
            sample=data_rows[:5], total=len(data_rows),
        )

    return render_template("patients/import.html")


def _load_import_tmp(token):
    """Return (path, payload) for a stashed import, or (None, None) if missing."""
    if not token.isalnum():
        return None, None
    tmp_path = os.path.join(_import_tmp_dir(), f"{token}.json")
    if not os.path.isfile(tmp_path):
        return None, None
    with open(tmp_path, encoding="utf-8") as fh:
        return tmp_path, json.load(fh)


@patients_bp.route("/import/map", methods=["POST"])
@module_required(MODULE)
def import_map():
    token = (request.form.get("token") or "").strip()
    tmp_path, payload = _load_import_tmp(token)
    # A freshly-stashed upload is a dict; older canonical payloads are a list.
    if payload is None or not isinstance(payload, dict):
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.bulk_import"))

    headers = payload["headers"]
    data_rows = payload["rows"]

    # Read the user's column choices: each field maps to a column index or "".
    mapping = {}
    for key, _required, _sample in import_fields():
        raw = (request.form.get(f"map_{key}") or "").strip()
        if raw == "":
            continue
        try:
            idx = int(raw)
        except ValueError:
            continue
        if 0 <= idx < len(headers):
            mapping[key] = idx

    missing = [k for k in REQUIRED_KEYS if k not in mapping]
    if missing:
        flash(t("import.map_required"), "danger")
        return render_template(
            "patients/import_map.html",
            token=token, headers=headers, filename=payload.get("filename", ""),
            fields=import_fields(), required_keys=REQUIRED_KEYS,
            guess=mapping or guess_mapping(headers),
            sample=data_rows[:5], total=len(data_rows), missing=missing,
        )

    rows = build_rows(data_rows, mapping)
    # Overwrite the stash with canonical rows for the confirm step.
    with open(tmp_path, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False, default=_json_cell)

    preview, valid = _analyze_rows(rows)
    return render_template(
        "patients/import_preview.html",
        token=token, preview=preview[:MAX_PREVIEW_ROWS],
        total=len(rows), valid=valid, invalid=len(rows) - valid,
        shown=min(len(rows), MAX_PREVIEW_ROWS),
        filename=payload.get("filename", ""),
    )


def _json_cell(value):
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)


@patients_bp.route("/import/confirm", methods=["POST"])
@module_required(MODULE)
def import_confirm():
    token = (request.form.get("token") or "").strip()
    tmp_path, rows = _load_import_tmp(token)
    # The confirm step expects canonical rows (a list) written by import_map.
    if rows is None or not isinstance(rows, list):
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.bulk_import"))

    result = _process_import(rows)
    db.session.commit()
    ActivityLog.record(
        "patient.import", user_id=current_user.id, entity="patient",
        detail=f"created={result['created']} skipped={len(result['errors'])}",
        ip_address=client_ip(),
    )
    db.session.commit()
    try:
        os.remove(tmp_path)
    except OSError:
        pass
    return render_template("patients/import_result.html", result=result)


@patients_bp.route("/import/template")
@module_required(MODULE)
def import_template():
    fmt = (request.args.get("fmt") or "xlsx").lower()
    if fmt == "csv":
        return Response(
            build_template_csv(),
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": "attachment; filename=patients_template.csv"
            },
        )
    return send_file(
        build_template_workbook(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="patients_template.xlsx",
    )


def _process_import(rows):
    """Create patients/families/parents from parsed rows.

    Families are de-duplicated by name within the import (and matched against
    existing families) so siblings land in the same family record.
    """
    from app.utils import patient_dedupe
    from app.utils.imports import family_key

    created = 0
    errors = []
    # Re-uploading a sheet used to double the register: every row was written,
    # every time. The index is loaded once and the rows are walked in memory,
    # like the history import — a real sheet is thousands of rows and asking
    # per row is what makes an import look hung.
    known = patient_dedupe.index()
    in_file = {}                 # the same child listed twice in one sheet
    skipped = []
    # Keyed by the *folded* household key rather than the name as typed.
    #
    # Grouping by the raw name is what put one family into three records:
    # Egyptian names run child → father → grandfather → family and a sheet
    # records as many as whoever filled it knew, so "محمود سعيد أحمد" and
    # "محمود سعيد" are the same father written twice — and "أحمد" / "احمد" are
    # one name to every human being and two strings to a computer.
    #
    # Existing families are indexed the same way, so a second import next
    # month joins the family the first one made instead of building a parallel
    # one beside it.
    family_cache = {}
    for existing in Family.query.all():
        family_cache.setdefault(family_key(existing.family_name), existing)
    phone_family = {}        # guardian phone -> Family (siblings share a phone)
    family_has_parent = set()  # id(family) that already carries a guardian
    next_number = patient_number_allocator()

    def get_named_family(name):
        """The family for this guardian name, matched on the folded key.

        The *stored* name stays the fullest version anybody wrote — the key is
        for matching, and "محمود سعيد أحمد" is a better thing to read on a
        screen than the two words used to group by.
        """
        key = family_key(name)
        if not key:
            return None
        family = family_cache.get(key)
        if family is None:
            family = Family(family_name=name)
            db.session.add(family)
            family_cache[key] = family
        elif len((name or "").split()) > len((family.family_name or "").split()):
            family.family_name = name
        return family

    def cell(key):
        val = row.get(key)
        if val in (None, ""):
            return None
        return str(val).strip() or None

    # ``no_autoflush`` keeps SQLAlchemy from flushing the growing batch of
    # pending patients on every read, which is what made large imports crawl.
    with db.session.no_autoflush:
        for offset, row in enumerate(rows):
            line = offset + 2  # account for the header row (1-based)
            name = (row.get("full_name") or "").strip()
            if not name:
                errors.append({"line": line, "reason": t("import.err_name")})
                continue

            gender = parse_gender(row.get("gender"))
            if gender is None:
                errors.append({"line": line, "reason": t("import.err_gender")})
                continue

            dob = parse_date(row.get("date_of_birth"))
            if dob is None:
                errors.append({"line": line, "reason": t("import.err_dob")})
                continue

            # Guardian name: explicit, or derived from the child's name (father
            # = the name after the child's first name) and flagged for review.
            given_parent = (row.get("parent_name") or "").strip()
            derived_parent = derive_guardian_name(name)
            parent_name = given_parent or derived_parent
            phone = normalize_phone(row.get("parent_phone"))
            fam_name = (row.get("family_name") or "").strip()

            # Resolve family so siblings group: explicit family name → shared
            # guardian phone → the (given/derived) guardian name.
            if fam_name:
                family = get_named_family(fam_name)
            elif phone and phone in phone_family:
                family = phone_family[phone]
            elif phone:
                family = get_named_family(parent_name or name)
                phone_family[phone] = family
            elif parent_name:
                family = get_named_family(parent_name)
            else:
                family = None

            # Already here — from an earlier upload or from earlier in this
            # same sheet. Left exactly as it is: the request was the
            # increment, and overwriting a file somebody has since corrected
            # by hand is worse than importing nothing.
            existing_id = patient_dedupe.match(row, dob, known)
            twin = (None if existing_id is not None
                    else patient_dedupe.match(row, dob, in_file))
            if existing_id is not None or twin is not None:
                skipped.append({"line": line, "name": name,
                                "patient_id": existing_id, "twin": twin})
                continue

            patient = Patient(
                patient_number=next_number(),
                reference_number=cell("reference_number"),
                family=family,
                full_name=name,
                full_name_en=cell("full_name_en"),
                date_of_birth=dob,
                gender=gender,
                national_id=cell("national_id"),
                blood_type=cell("blood_type"),
                allergies=cell("allergies"),
                chronic_diseases=cell("chronic_diseases"),
                notes=cell("notes"),
                is_active=True,
            )
            db.session.add(patient)

            # One guardian per family — skip if the family already has one
            # (existing record or an earlier sibling in this import).
            if (parent_name and family is not None
                    and id(family) not in family_has_parent and not family.parents):
                relation = (row.get("parent_relation") or "father").strip().lower()
                category = (row.get("client_category") or "normal").strip().lower()
                db.session.add(Parent(
                    family=family,
                    relation=relation if Parent.valid_relation(relation) else "father",
                    full_name=parent_name,
                    auto_named=not given_parent,
                    phone=cell("parent_phone"),
                    phone_alt=cell("parent_phone_alt"),
                    national_id=cell("parent_national_id"),
                    email=cell("parent_email"),
                    occupation=cell("parent_occupation"),
                    nationality=cell("parent_nationality"),
                    address=cell("parent_address"),
                    client_category=category if Parent.valid_category(category) else "normal",
                ))
                family_has_parent.add(id(family))

            # Remembered by line, not by id: the patients are still pending
            # and flushing to get one would undo the batching that keeps a
            # ten-thousand-row import from crawling.
            for key in patient_dedupe.row_keys(row, dob):
                in_file.setdefault(key, line)
            created += 1

    return {"created": created, "errors": errors, "total": len(rows),
            "skipped": skipped}


# ==================================================== history import ========
# Bringing a clinic's old case history across from the program it used before.
# Same wizard as the patient import above — upload, map, preview, commit — so
# somebody who has done one recognises the other. The differences are all
# consequences of scale and of what the file cannot say:
#
#   * it attaches to patients that must already exist (an old services export
#     carries no date of birth, gender or phone, so it cannot create anybody);
#   * every lookup is done in bulk, because ten thousand rows resolved one at a
#     time is ten thousand round trips;
#   * and a second upload is *compared*, not appended.
def _history_cell(value):
    """JSON-safe cell that keeps the time of day.

    The patient import's serialiser formats dates as ``%Y-%m-%d``, which is
    right there and wrong here: the time is part of what tells two services on
    the same day apart, and dropping it makes 80 rows of a real export look
    like duplicates of each other.
    """
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dtime):
        return value.strftime("%H:%M:%S")
    return str(value)


def _history_ready():
    """Whether there are any patients for the history to attach to."""
    return Patient.query.limit(1).count() > 0


@patients_bp.route("/import/history", methods=["GET", "POST"])
@module_required(MODULE)
def history_import():
    from app.utils.history_import import guess_mapping, summary_columns
    from app.utils.history_import import fields as history_fields

    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash(t("import.no_file"), "danger")
            return redirect(url_for("patients.history_import"))
        if not allowed_import_file(file.filename):
            flash(t("import.bad_format"), "danger")
            return redirect(url_for("patients.history_import"))

        headers, data_rows, error = read_matrix(file)
        if error == "unreadable":
            flash(t("import.unreadable"), "danger")
            return redirect(url_for("patients.history_import"))
        if error or not data_rows:
            flash(t("import.empty"), "warning")
            return redirect(url_for("patients.history_import"))

        # Parsed once and stashed: the mapping screen, the preview and the
        # commit all read this instead of re-reading the workbook three times.
        token = uuid.uuid4().hex
        with open(os.path.join(_import_tmp_dir(), f"{token}.json"), "w",
                  encoding="utf-8") as fh:
            json.dump({"headers": headers, "rows": data_rows,
                       "filename": file.filename},
                      fh, ensure_ascii=False, default=_history_cell)

        return render_template(
            "patients/history_map.html", token=token, headers=headers,
            filename=file.filename, fields=history_fields(),
            guess=guess_mapping(headers), sample=data_rows[:5],
            total=len(data_rows),
            # Trailing columns that are a summary block rather than data — the
            # real export ends with "من تاريخ / إلى تاريخ / عدد الخدمات" laid
            # out like columns inside the same sheet.
            summary=sorted(summary_columns(headers, data_rows)))

    return render_template("patients/history_import.html",
                           ready=_history_ready())


@patients_bp.route("/import/history/map", methods=["POST"])
@module_required(MODULE)
def history_import_map():
    from app.utils.export import parse_date
    from app.utils.history_import import build_rows
    from app.utils.history_match import (classify, date_span, distinct_values,
                                         missing_patient_codes)

    token = (request.form.get("token") or "").strip()
    tmp_path, payload = _load_import_tmp(token)
    if payload is None:
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.history_import"))

    # Submitted the first time through and stored; coming back to change the
    # date range reuses it instead of asking again.
    mapping = _submitted_mapping(payload)
    if not mapping:
        flash(t("history_import.need_columns"), "danger")
        return redirect(url_for("patients.history_import"))

    records = build_rows(payload["rows"], mapping)
    # What the file covers is read off the file. Defaulting to all of it is the
    # case that needs no thought — a clinic leaving its old program wants
    # everything, and making it say so is a step that exists only to be got
    # wrong.
    span = date_span(records)
    links = payload.get("links") or {}
    start = parse_date(request.form.get("from"))
    end = parse_date(request.form.get("to"))
    records, counts = classify(records, start=start, end=end)

    with open(tmp_path, "w", encoding="utf-8") as fh:
        json.dump({**payload, "mapping": mapping}, fh, ensure_ascii=False,
                  default=_history_cell)

    return render_template(
        "patients/history_preview.html", token=token, counts=counts,
        total=len(records), filename=payload.get("filename"),
        rows=[r for r in records
              if r["_state"] != "out_of_range"][:MAX_PREVIEW_ROWS],
        span=span, f_from=request.form.get("from", ""),
        f_to=request.form.get("to", ""), links=links,
        linked=sum(1 for v in links.values() if str(v).startswith("brand:")),
        missing_patients=missing_patient_codes(records),
        services=distinct_values(records, "service_name"),
        categories=distinct_values(records, "client_category"),
        doctors=distinct_values(records, "doctor_name"),
        # Named before anything is written. Creating users is the one part of
        # this import that adds people to the clinic rather than history, so it
        # is the part that must not happen quietly.
        new_doctors=_doctors_to_create(payload))


@patients_bp.route("/import/history/commit", methods=["POST"])
@module_required(MODULE)
def history_import_commit():
    from app.models import ImportBatch, ImportedService
    from app.utils.dose_infer import number_doses
    from app.utils.history_import import build_rows
    from app.utils.history_match import CHANGED, NEW, classify, doctor_key

    token = (request.form.get("token") or "").strip()
    tmp_path, payload = _load_import_tmp(token)
    if payload is None or not payload.get("mapping"):
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.history_import"))

    from app.utils.export import parse_date

    start = parse_date(request.form.get("from"))
    end = parse_date(request.form.get("to"))
    records = build_rows(payload["rows"], payload["mapping"])
    # The same range the preview was shown for, carried on the form — otherwise
    # somebody narrows the range, sees 400 rows, and imports ten thousand.
    records, counts = classify(records, start=start, end=end)

    batch = ImportBatch(kind="history", filename=payload.get("filename"),
                        created_by=current_user.id, rows_total=len(records),
                        # Asked on the preview, stored on the batch. Off unless
                        # somebody ticked it: replaying ten years as revenue
                        # counts a decade twice.
                        count_money=request.form.get("count_money") == "1")
    db.session.add(batch)
    db.session.flush()

    # Written with one bulk insert and one commit. Adding rows one at a time
    # here costs a round trip each — and on SQLite, committing per row costs a
    # disk sync each, which is the difference between a second and an hour.
    update_changed = request.form.get("update_changed") == "1"

    # What the clinic confirmed on the linking screen: "this name is that
    # brand". Resolved once into {name: (brand_id, vaccine_id)} rather than
    # per row — 9,908 rows carry 27 names.
    links = _resolved_links(payload.get("links") or {})
    # The same screen also links the rows that are not vaccines at all — كشف
    # and إستشارة are 7,476 of the real file — onto the services the clinic
    # already has, so imported history lands in the same catalogue as today's.
    service_links = _resolved_services(payload.get("links") or {})
    # The same "map onto what exists" rule, for the two columns that were
    # previously stored as text and therefore invisible to every report.
    doctor_links = _resolved_doctors(payload.get("doctor_links") or {})
    # And the doctors the file names that this clinic has no user for. Created
    # here rather than on the linking screen, so backing out at the preview
    # leaves nothing behind: users appear only alongside the history that
    # needed them.
    doctor_links.update(_created_doctors(payload))
    category_links = payload.get("category_links") or {}

    pending = []
    for record in records:
        if record["_state"] != NEW and not (update_changed
                                            and record["_state"] == CHANGED):
            continue
        if record["_state"] == CHANGED:
            ImportedService.query.filter_by(source_key=record["_key"]).delete()
        brand_id, vaccine_id = links.get(record["service_name"], (None, None))
        pending.append({
            "brand_id": brand_id, "vaccine_id": vaccine_id,
            "service_id": service_links.get(record["service_name"]),
            # Without this a decade of a doctor's work sits outside the
            # commission reports, the doctor filter and the statements — every
            # one of which joins on doctor_id.
            "doctor_id": doctor_links.get(doctor_key(record)),
            "batch_id": batch.id, "patient_id": record["_patient_id"],
            "service_date": record["service_date"],
            "service_time": record["service_time"],
            "source_name": record["service_name"][:255],
            "quantity": record["quantity"], "price": record["price"],
            "doctor_share": record["doctor_share"],
            "paid_cash": record["paid_cash"],
            "paid_company": record["paid_company"],
            "client_category": (category_links.get(record["client_category"])
                                or (record["client_category"] or "")[:30]
                                or None),
            "source_key": record["_key"], "source_row": record["source_row"][:40],
            "notes": record["notes"] or None,
        })

    # Dose numbers last, over the whole set at once — the number depends on the
    # patient's other doses of the *same vaccine*, so it cannot be worked out a
    # row at a time. Numbering per brand instead would restart the course every
    # time a clinic switched product, and the schedule would then chase a child
    # for doses they have already had.
    number_doses(pending)
    for row in pending:
        row["vaccine_brand_id"] = row.pop("brand_id")
        row.pop("vaccine_id", None)

    if pending:
        db.session.bulk_insert_mappings(ImportedService, pending)

    # And the vaccinations become real vaccination records. Without this the
    # import is a wall of text beside the file that already ignores it: the
    # schedule counts PatientVaccine rows, so a child whose whole course was
    # imported would be chased by the reminder screen for doses they have had.
    made = _record_imported_doses(pending, batch.id)

    batch.rows_added = len(pending)
    batch.notes = f"vaccinations: {made}" if made else None
    batch.rows_skipped = (counts["same"] + counts["out_of_range"]
                          + (0 if update_changed else counts["changed"]))
    batch.rows_rejected = counts["rejected"]
    ActivityLog.record("history.import", user_id=current_user.id,
                       entity="import_batch", entity_id=batch.id,
                       detail=f"{len(pending)} rows from {payload.get('filename')}",
                       ip_address=client_ip())
    db.session.commit()

    try:
        os.remove(tmp_path)
    except OSError:
        pass

    flash(t("history_import.done").replace("{n}", str(len(pending))), "success")
    return redirect(url_for("patients.history_import"))


@patients_bp.route("/import/history/missing", methods=["POST"])
@module_required(MODULE)
def history_missing_patients():
    """The unknown patients in the file, as a patient-import sheet to fill in.

    The honest answer to "what happens to a child in the file who is not in the
    program yet" has two halves, and only the first was built.

    **Their history is not lost.** Nothing is imported for them, and nothing is
    thrown away either: the rows stay in the file, and a re-upload after the
    patients exist picks them up, because a second upload compares instead of
    appending. So the fix is always available and never urgent.

    **But "import those patients first" was a sentence, not a path.** A
    services export has no date of birth, gender or phone — so the patients
    cannot be created from it, and telling somebody to go and type eighty-seven
    children by hand is how a clinic decides to skip its own history.

    This meets it halfway: the file the clinic uploads to the *patient* import,
    already carrying the code and the name from their own export, with the
    columns only they can fill left blank. **The date of birth is deliberately
    one of those.** It drives every vaccination due date and every growth
    centile, so a guessed one would be a wrong dose date for a real child —
    the one thing worth more than the convenience.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.utils.history_import import build_rows
    from app.utils.history_match import classify, missing_patient_codes
    from app.utils.imports import IMPORT_COLUMNS

    token = (request.form.get("token") or "").strip()
    _tmp_path, payload = _load_import_tmp(token)
    if payload is None or not payload.get("mapping"):
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.history_import"))

    records, _counts = classify(build_rows(payload["rows"], payload["mapping"]))
    missing = missing_patient_codes(records, limit=None)
    if not missing:
        flash(t("history_import.no_missing"), "info")
        return redirect(url_for("patients.history_import"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"
    headers = [c[0] for c in IMPORT_COLUMNS]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="198754")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # The old program's code goes in `reference_number`, which is the column
    # the history import matches on afterwards. Putting it anywhere else would
    # produce patients the second upload still cannot find.
    ref_at = headers.index("reference_number")
    name_at = headers.index("full_name")
    for entry in missing:
        row = [""] * len(headers)
        row[ref_at] = entry["code"]
        row[name_at] = entry["name"]
        ws.append(row)

    info = wb.create_sheet("تعليمات")
    for line in (t("history_import.missing_sheet_1"),
                 t("history_import.missing_sheet_2"),
                 t("history_import.missing_sheet_3")):
        info.append([line])
    info.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="missing_patients.xlsx")


@patients_bp.route("/import/history/template")
@module_required(MODULE)
def history_import_template():
    """A template with the expected columns and a sample row.

    A convenience, not a requirement: the real export this was built against
    maps 16 of its 17 columns with nobody renaming anything, and that has to
    stay the normal case. The template is for a clinic whose program exports
    something unrecognisable, or one typing its history in by hand.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    from app.utils.history_import import fields as history_fields

    wb = Workbook()
    ws = wb.active
    ws.title = "history"
    columns = history_fields()
    ws.append([label for _k, _r, label in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEDD")
    ws.append(["1", "2024-03-15", "10:30", "1043", "أحمد محمد", "1",
               "د. سارة أحمد", "كشف", "", "الكشف", "نقدي", "200", "80",
               "200", "0", "1", ""])
    for index, _col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="history_template.xlsx")


@patients_bp.route("/import/history/link", methods=["POST"])
@module_required(MODULE)
def history_import_link():
    """Confirm what each of the file's service names actually is.

    The step that makes the vaccinations real. 9,908 rows carry 27 distinct
    names, so this screen is 27 rows: each with the catalogue's best guess and
    a confidence, and the clinic changes what it disagrees with.

    Nothing is matched automatically into the records. A matcher that wrote its
    own guesses into ten years of vaccination history would be one nobody could
    trust, and the case it gets wrong is a child recorded as having had a
    vaccine they did not.
    """
    from app.utils.history_import import build_rows
    from app.utils.history_match import distinct_values, doctor_entries
    from app.utils.vaccine_match import suggest_all

    token = (request.form.get("token") or "").strip()
    tmp_path, payload = _load_import_tmp(token)
    if payload is None:
        flash(t("import.session_expired"), "warning")
        return redirect(url_for("patients.history_import"))

    # The column mapping arrives here the first time through, from the mapping
    # screen, and is stored — coming back to change a link must not send
    # somebody through the columns again.
    mapping = _submitted_mapping(payload)
    if not mapping:
        flash(t("history_import.need_columns"), "danger")
        return redirect(url_for("patients.history_import"))
    if mapping != (payload.get("mapping") or {}):
        payload["mapping"] = mapping
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, default=_history_cell)

    records = build_rows(payload["rows"], mapping)
    names = distinct_values(records, "service_name")
    # Grouped by the source's own doctor code where the file has one: a name is
    # typed several ways across ten years, and one row per spelling would offer
    # the clinic the same person three times — then create three users.
    doctors = doctor_entries(records)
    categories = distinct_values(records, "client_category")

    # Saving the screen: store what the clinic chose and go on to the preview.
    if request.form.get("confirm") == "1":
        links = {}
        for index, row in enumerate(names):
            choice = (request.form.get(f"link_{index}") or "").strip()
            if choice:
                links[row["value"]] = choice
        doctor_links = {}
        # Keyed by the doctor's *key* — the code when the file has one — and
        # the name is stored beside it so the preview can say who is about to
        # be created without re-reading the file.
        doctor_names = {}
        for index, row in enumerate(doctors):
            choice = (request.form.get(f"doc_{index}") or "").strip()
            if choice:
                doctor_links[row["key"]] = choice
                doctor_names[row["key"]] = row["value"]
        category_links = {}
        for index, row in enumerate(categories):
            choice = (request.form.get(f"cat_{index}") or "").strip()
            if choice:
                category_links[row["value"]] = choice
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump({**payload, "links": links,
                       "doctor_links": doctor_links,
                       "doctor_names": doctor_names,
                       "category_links": category_links},
                      fh, ensure_ascii=False, default=_history_cell)
        return history_import_map()

    # Every column that points at something in the program is matched against
    # what exists first — the file's doctor and its "التعاقد" as much as its
    # vaccine names. Creating is never the default.
    from app.utils.history_match import normalise_arabic
    from app.utils.name_match import suggest_doctors, suggest_services
    from app.utils.client_categories import active_categories, ensure_seeded

    ensure_seeded()
    category_rows = active_categories()
    by_category = {normalise_arabic(c.display_name("ar")): c.key
                   for c in category_rows}
    by_category.update({normalise_arabic(c.key): c.key for c in category_rows})

    # Two catalogues per name, not one. 7,476 of the real file's rows are كشف
    # and إستشارة — services the clinic already has, priced and commissioned —
    # and offering only "a plain service" threw that away: the row kept its
    # text and pointed at nothing, so it never reached a revenue report.
    name_values = [row["value"] for row in names]
    vaccine_hits = suggest_all(name_values)
    service_hits = suggest_services(name_values)

    return render_template(
        "patients/history_link.html", token=token, names=names,
        suggestions=vaccine_hits, service_suggestions=service_hits,
        best_link=_best_link(name_values, vaccine_hits, service_hits),
        saved=payload.get("links") or {},
        filename=payload.get("filename"),
        # Every brand and every service, for the rows the matcher could not
        # place. Loaded once — a select per row, not a query per row.
        brands=_brand_choices(), services=_service_choices(),
        doctors=doctors, doctor_choices=_doctor_choices(),
        doctor_hits=suggest_doctors([row["value"] for row in doctors]),
        saved_doctors=payload.get("doctor_links") or {},
        # "Create an inactive user" is the last option, never the default —
        # matching what exists comes first, and nothing is written until the
        # preview has named who is about to be created.
        create_doctor_choice=CREATE_DOCTOR,
        categories=categories, category_choices=category_rows,
        category_guess={row["value"]: by_category.get(
            normalise_arabic(row["value"])) for row in categories},
        saved_categories=payload.get("category_links") or {})


def _best_link(values, vaccine_hits, service_hits):
    """``{name: (choice, confidence)}`` — the one proposal per row.

    A name is scored against both catalogues, and the higher score wins. The
    tie goes to the **vaccine**, because a vaccination row carries a dose
    number and a course: linking it to the "vaccination fee" service instead
    would price it correctly and still leave the child's schedule unaware the
    dose happened.
    """
    out = {}
    for value in values:
        vaccine = (vaccine_hits.get(value) or [None])[0]
        service = (service_hits.get(value) or [None])[0]
        if vaccine and (not service or vaccine["score"] >= service["score"]):
            out[value] = (f"brand:{vaccine['brand_id']}", vaccine["confidence"])
        elif service:
            out[value] = (f"service:{service['service_id']}",
                          service["confidence"])
    return out


def _service_choices():
    """``[(service_id, name)]`` for the "which service is this?" dropdown."""
    from app.models import Service

    lang = getattr(g, "lang", "ar")
    rows = Service.query.filter(Service.is_active.is_(True)).all()
    return sorted(((s.id, s.display_name(lang)) for s in rows),
                  key=lambda row: row[1])


def _doctor_choices():
    """``[(user_id, name)]`` for the "who is this doctor?" dropdown."""
    from app.utils.appointments import list_doctors

    lang = getattr(g, "lang", "ar")
    return [(u.id, u.display_name(lang)) for u in list_doctors()]


def _submitted_mapping(payload):
    """The column mapping for this upload: from the form, else what was stored.

    Required columns are checked here rather than in each step, so a wizard
    entered halfway — somebody re-posting to change one link — cannot proceed
    on a mapping that never existed.
    """
    mapping = {}
    for key in request.form:
        if not key.startswith("col_"):
            continue
        value = (request.form.get(key) or "").strip()
        if value != "":
            mapping[key[4:]] = int(value)
    if not mapping:
        mapping = {k: int(v) for k, v in (payload.get("mapping") or {}).items()}
    if any(k not in mapping
           for k in ("service_date", "patient_code", "service_name")):
        return {}
    return mapping


def _brand_choices():
    """``[(brand_id, label)]`` for the "which vaccine is this?" dropdown."""
    from app.models import Vaccine, VaccineBrand

    vaccines = {v.id: v for v in Vaccine.query.all()}
    out = []
    for brand in VaccineBrand.query.order_by(VaccineBrand.name).all():
        vaccine = vaccines.get(brand.vaccine_id)
        label = f"{vaccine.name_ar} — {brand.name}" if vaccine else brand.name
        out.append((brand.id, label))
    return sorted(out, key=lambda row: row[1])


def _resolved_services(links):
    """``{service name in the file: service id}`` from the confirmed choices.

    Only ids that still exist are kept, for the same reason as the doctors: a
    mapping saved against a service somebody has since deleted would write a
    dangling reference into ten years of history.
    """
    from app.models import Service

    wanted = {}
    for name, choice in links.items():
        choice = str(choice or "")
        if not choice.startswith("service:"):
            continue
        try:
            wanted[name] = int(choice.split(":", 1)[1])
        except ValueError:
            continue
    if not wanted:
        return {}
    real = {s.id for s in Service.query.filter(
        Service.id.in_(set(wanted.values()))).all()}
    return {name: sid for name, sid in wanted.items() if sid in real}


def _resolved_links(links):
    """``{service name: (brand_id, vaccine_id)}`` from the confirmed choices.

    The vaccine comes from the brand rather than being stored beside it,
    because the dose number is counted per *vaccine* and a brand that was
    re-filed under a different vaccine must not leave old links pointing at
    the wrong course.
    """
    from app.models import VaccineBrand

    wanted = {}
    for name, choice in links.items():
        if not str(choice).startswith("brand:"):
            continue
        try:
            wanted[name] = int(str(choice).split(":", 1)[1])
        except ValueError:
            continue
    if not wanted:
        return {}
    brands = {b.id: b.vaccine_id for b in VaccineBrand.query.filter(
        VaccineBrand.id.in_(set(wanted.values()))).all()}
    return {name: (brand_id, brands.get(brand_id))
            for name, brand_id in wanted.items() if brand_id in brands}


def _record_imported_doses(rows, batch_id):
    """Turn the linked import rows into real vaccination records.

    The schedule, the reminders and the certificate all read ``PatientVaccine``.
    An imported dose that stays outside that table is history the program can
    show and cannot *use* — the reminder screen would still chase the child for
    a dose they had in 2023.

    Three things this deliberately does not do:

    * **It does not touch the fridge.** ``inventory_id`` stays empty: the vial
      was used years ago at another program, and deducting it now would invent
      a stock movement that never happened here.
    * **It does not overwrite what the clinic already recorded.** A dose the
      nurse entered by hand outranks one inferred from dates, so an existing
      record for the same patient, vaccine and dose number is left alone.
    * **It does not hide where it came from.** ``import_batch_id`` marks these
      as the doses whose numbering was *inferred* rather than observed — which
      is exactly the set a doctor may need to correct, and what makes an import
      undoable without touching anything typed since.
    """
    from app.models import PatientVaccine

    wanted = [r for r in rows if r.get("vaccine_brand_id") and r.get("dose_number")]
    if not wanted:
        return 0

    # One query for everything already on file, rather than one per dose.
    patient_ids = {r["patient_id"] for r in wanted}
    existing = {
        (pid, vid, dose) for pid, vid, dose in
        db.session.query(PatientVaccine.patient_id, PatientVaccine.vaccine_id,
                         PatientVaccine.dose_number)
        .filter(PatientVaccine.patient_id.in_(patient_ids),
                PatientVaccine.event_type == "given").all()}

    from app.models import VaccineBrand
    brand_vaccine = {
        b.id: b.vaccine_id for b in VaccineBrand.query.filter(
            VaccineBrand.id.in_({r["vaccine_brand_id"] for r in wanted})).all()}

    doses = []
    for row in wanted:
        vaccine_id = brand_vaccine.get(row["vaccine_brand_id"])
        if not vaccine_id:
            continue
        key = (row["patient_id"], vaccine_id, row["dose_number"])
        if key in existing:
            continue
        existing.add(key)
        doses.append({
            "patient_id": row["patient_id"], "vaccine_id": vaccine_id,
            "brand_id": row["vaccine_brand_id"],
            "dose_number": row["dose_number"],
            "given_date": row["service_date"], "event_type": "given",
            "given_outside": False, "import_batch_id": batch_id,
        })
    if doses:
        db.session.bulk_insert_mappings(PatientVaccine, doses)
    return len(doses)


@patients_bp.route("/import/history/batches")
@module_required(MODULE)
def history_batches():
    """Every history import this clinic has run, newest first."""
    from app.models import ImportBatch

    return render_template(
        "patients/history_batches.html",
        batches=(ImportBatch.query.filter_by(kind="history")
                 .order_by(ImportBatch.id.desc()).limit(50).all()))


@patients_bp.route("/import/history/batches/<int:batch_id>/undo",
                   methods=["POST"])
@module_required(MODULE)
def history_batch_undo(batch_id):
    """Take back one import — and nothing else.

    Ten thousand rows written against real data needs a way back, and the only
    kind worth having is one that is *exact*. Every row this import created
    carries its batch, so undoing removes what it added and leaves untouched
    everything the clinic has entered or corrected since.

    Two things are deliberately kept:

    * **A dose the clinic has since corrected.** If a doctor changed the number,
      the date, or marked it given elsewhere, that is their record now — not the
      import's — and deleting it would throw away the review the whole import
      was built to invite.
    * **The batch row itself.** "This import was undone, by whom, when" is part
      of the history of the file, and a clinic asking six months later why a
      decade of vaccinations is missing deserves an answer.
    """
    from app.models import ImportBatch, ImportedService, PatientVaccine

    batch = db.get_or_404(ImportBatch, batch_id)
    if batch.rows_added == 0 and batch.notes == "undone":
        flash(t("history_import.already_undone"), "info")
        return redirect(url_for("patients.history_batches"))

    # Doses the clinic has corrected since are theirs, not the import's.
    kept = 0
    doses = PatientVaccine.query.filter_by(import_batch_id=batch.id).all()
    for dose in doses:
        if dose.given_outside or dose.outside_place or dose.lot_number:
            dose.import_batch_id = None
            kept += 1
        else:
            db.session.delete(dose)

    removed = ImportedService.query.filter_by(batch_id=batch.id).delete(
        synchronize_session=False)

    batch.rows_added = 0
    batch.notes = "undone"
    ActivityLog.record("history.import.undo", user_id=current_user.id,
                       entity="import_batch", entity_id=batch.id,
                       detail=f"-{removed} rows, {len(doses) - kept} doses, "
                              f"{kept} kept",
                       ip_address=client_ip())
    db.session.commit()
    flash(t("history_import.undone").replace("{n}", str(removed)), "success")
    return redirect(url_for("patients.history_batches"))


@patients_bp.route("/imported/<int:row_id>/edit", methods=["POST"])
@module_required(MODULE)
def imported_service_edit(row_id):
    """Correct one line of imported history.

    Asked for as "add these services to the patient files and edit them there
    one at a time". Ten years of somebody else's data will have wrong dates,
    wrong prices and names that mean nothing here, and a clinic that cannot fix
    a line in its own file does not trust the file.
    """
    from app.models import ImportedService

    row = db.get_or_404(ImportedService, row_id)
    name = (request.form.get("source_name") or "").strip()
    if name:
        row.source_name = name[:255]
    when = (request.form.get("service_date") or "").strip()
    if when:
        try:
            row.service_date = datetime.strptime(when, "%Y-%m-%d").date()
        except ValueError:
            pass
    price = request.form.get("price", type=float)
    if price is not None and price >= 0:
        row.price = price
    row.notes = (request.form.get("notes") or "").strip() or None

    ActivityLog.record("history.row.edit", user_id=current_user.id,
                       entity="patient", entity_id=row.patient_id,
                       detail=f"{row.source_name} {row.service_date}",
                       ip_address=client_ip())
    db.session.commit()
    flash(t("history_import.row_saved"), "success")
    return redirect(url_for("patients.view", patient_id=row.patient_id)
                    + "#history")


@patients_bp.route("/imported/<int:row_id>/delete", methods=["POST"])
@module_required(MODULE)
def imported_service_delete(row_id):
    """Remove one line of imported history.

    Deleting the whole import is a different button on a different screen. This
    is for the single row that should never have been there — a duplicate in
    the old program, a service billed to the wrong child years ago.
    """
    from app.models import ImportedService

    row = db.get_or_404(ImportedService, row_id)
    patient_id = row.patient_id
    ActivityLog.record("history.row.delete", user_id=current_user.id,
                       entity="patient", entity_id=patient_id,
                       detail=f"{row.source_name} {row.service_date}",
                       ip_address=client_ip())
    db.session.delete(row)
    db.session.commit()
    flash(t("history_import.row_removed"), "info")
    return redirect(url_for("patients.view", patient_id=patient_id)
                    + "#history")


def _created_doctors(payload):
    """Make the users the clinic asked for, and return ``{key: user id}``.

    Written into the same transaction as the rows, so a clinic that abandons
    the import does not leave a list of half-created doctors behind.
    """
    from app.utils.import_doctors import create_all

    names = payload.get("doctor_names") or {}
    wanted = {key: (names.get(key) or "").strip()
              for key, choice in (payload.get("doctor_links") or {}).items()
              if str(choice) == CREATE_DOCTOR}
    wanted = {k: v for k, v in wanted.items() if v}
    if not wanted:
        return {}

    made = create_all(sorted(set(wanted.values())))
    for name, user in made.items():
        ActivityLog.record("history.doctor.create", user_id=current_user.id,
                           entity="user", entity_id=user.id,
                           detail=name, ip_address=client_ip())
    return {key: made[name].id for key, name in wanted.items() if name in made}


def _doctors_to_create(payload):
    """The names the clinic asked to be made into users, in file order."""
    names = payload.get("doctor_names") or {}
    return [names.get(key) or key
            for key, choice in (payload.get("doctor_links") or {}).items()
            if str(choice) == CREATE_DOCTOR]


@patients_bp.route("/<int:patient_id>/siblings/link", methods=["POST"])
@module_required(MODULE)
def sibling_link(patient_id):
    """Attach an existing patient to this child's family.

    The action that was missing. Creating a second file for a child who
    already has one is the expensive mistake here — it splits a vaccination
    history in half, and the half the next receptionist finds is the empty one.

    A child who is already filed with **another** family is refused rather
    than moved: that is a merge, and a merge is not a one-click button on a
    list of suggestions.
    """
    from app.models import Family

    patient = db.get_or_404(Patient, patient_id)
    other = db.session.get(Patient, request.form.get("sibling_id", type=int))
    if other is None or other.id == patient.id:
        flash(t("common.not_found"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#family")
    # A child who already has a family used to be refused here, on the
    # reasoning that joining two households is a merge of two sets of parents
    # and not a one-click button. A clinic showed me why that was wrong, with
    # names: "عمر محمد السيد خفاجة" and "ميرال محمد السيد خفاجه" arrived from
    # one import in two families, because the import split them. The screen
    # then suggested each to the other and refused every attempt to act on the
    # suggestion — and renaming either family changed nothing, because the
    # problem was never the name.
    #
    # The commonest case is not two real households. It is one household the
    # program itself divided, and refusing the merge left the only way out as
    # deleting a family by hand. So it merges, and says what it did.
    merged_from = None
    if other.family_id and other.family_id != patient.family_id:
        merged_from = other.family

    # This child may have no family row yet — the commonest case of all, since
    # a family is created the first time somebody records a parent.
    if not patient.family_id:
        family = Family(family_name=patient.full_name)
        db.session.add(family)
        db.session.flush()
        patient.family_id = family.id

    other.family_id = patient.family_id
    other.family_auto = False          # a person looked at both files

    # If that emptied the family they came from, carry its guardians across
    # and remove the shell. Leaving it behind would keep the parents' phone
    # numbers attached to a household with no children in it — invisible on
    # every screen, and still turning up in searches.
    db.session.flush()          # so the old family's child list is current
    if merged_from is not None and not merged_from.patients:
        known = {(p.phone or "").strip() for p in (patient.family.parents or [])
                 if (p.phone or "").strip()}
        for parent in list(merged_from.parents):
            # Moved through the collections, not by reassigning the id.
            # ``Family.parents`` cascades delete-orphan, so a parent that is
            # still in the old family's list when it is deleted goes with it —
            # setting `family_id` alone silently loses the guardian's phone
            # number, which is the one thing the merge was supposed to keep.
            merged_from.parents.remove(parent)
            if parent.phone and (parent.phone or "").strip() in known:
                db.session.delete(parent)   # the same guardian, twice
            else:
                patient.family.parents.append(parent)
        db.session.delete(merged_from)
    ActivityLog.record("patient.sibling.link", user_id=current_user.id,
                       entity="patient", entity_id=patient.id,
                       detail=f"{other.patient_number} → family {patient.family_id}",
                       ip_address=client_ip())
    db.session.commit()
    name = other.display_name(getattr(g, "lang", "ar"))
    flash(t("siblings.merged" if merged_from is not None else "siblings.linked")
          .replace("{name}", name), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#family")


@patients_bp.route("/<int:patient_id>/siblings/unlink", methods=["POST"])
@module_required(MODULE)
def sibling_unlink(patient_id):
    """Take a child back out of the family — because linking the wrong one is
    a thing that happens, and an action with no way back is one people avoid
    using at all."""
    patient = db.get_or_404(Patient, patient_id)
    other = db.session.get(Patient, request.form.get("sibling_id", type=int))
    if other is None or other.family_id != patient.family_id:
        flash(t("common.not_found"), "danger")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#family")

    other.family_id = None
    # Clear the marker too: a child linked again by hand tomorrow must not
    # still be labelled as the program's guess.
    other.family_auto = False
    ActivityLog.record("patient.sibling.unlink", user_id=current_user.id,
                       entity="patient", entity_id=patient.id,
                       detail=other.patient_number, ip_address=client_ip())
    db.session.commit()
    flash(t("siblings.unlinked"), "info")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#family")


@patients_bp.route("/<int:patient_id>/siblings/search")
@module_required(MODULE)
def sibling_search(patient_id):
    """JSON: patients who could be linked as a sibling of this one."""
    from flask import jsonify

    from app.utils.patients import apply_patient_search

    patient = db.get_or_404(Patient, patient_id)
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"patients": []})
    rows = (apply_patient_search(
        Patient.query.filter(Patient.is_active.is_(True),
                             Patient.id != patient.id), q)
        .order_by(Patient.full_name).limit(15).all())
    lang = getattr(g, "lang", "ar")
    mine = ((patient.family.family_name or "").strip()
            if patient.family else "")
    out = []
    for p in rows:
        theirs = ((p.family.family_name or "").strip()
                  if p.family_id and p.family else "")
        out.append({
            "id": p.id, "full_name": p.display_name(lang),
            "patient_number": p.patient_number,
            # Said on the row rather than found out after pressing: linking a
            # child who already has a family joins two households, which is
            # not what the button looks like it does.
            "in_family": bool(p.family_id and p.family_id != patient.family_id),
            "family_name": theirs or None,
            # …unless both records carry the same typed name, which is one
            # household the program divided. Saying "another family" there
            # reads as the program being wrong.
            "same_name": bool(theirs and mine and theirs == mine
                              and p.family_id != patient.family_id)})
    return jsonify({"patients": out})


def _resolved_doctors(links):
    """``{name in the file: user id}`` from the confirmed choices.

    Only ids that still exist are kept: a mapping saved against a user who has
    since been deleted would write a dangling reference into ten years of
    history, and a dangling doctor is worse than none — the reports would count
    the work and be unable to say whose it was.
    """
    from app.models import User

    wanted = {}
    for name, choice in links.items():
        choice = str(choice or "").strip()
        if not choice.isdigit():
            continue
        wanted[name] = int(choice)
    if not wanted:
        return {}
    real = {u.id for u in User.query.filter(
        User.id.in_(set(wanted.values()))).all()}
    return {name: uid for name, uid in wanted.items() if uid in real}


@patients_bp.route("/families/<int:family_id>/delete", methods=["POST"])
@module_required(MODULE)
def family_delete(family_id):
    """Remove a family, leaving its children as unattached files.

    Needed because the history import links by the father's name, and a run
    over messy data produces families that should never have existed —
    somebody else's children under one man's name. Until now the only way out
    was to unlink each child one at a time and leave the empty shell behind.

    The children are kept, always. A family is a grouping, not a container:
    deleting the grouping must never take a patient's file with it, so the
    rows are detached first and the empty family is what gets removed.
    """
    from app.models import Family

    family = db.get_or_404(Family, family_id)
    freed = list(family.patients)
    for patient in freed:
        patient.family_id = None
    # Parents belong to the family rather than to any one child, so they go
    # with it — there is nothing left for them to describe.
    for parent in list(getattr(family, "parents", []) or []):
        db.session.delete(parent)
    db.session.delete(family)
    ActivityLog.record("patient.family.delete", user_id=current_user.id,
                       entity="family", entity_id=family_id,
                       detail=f"{len(freed)} freed", ip_address=client_ip())
    db.session.commit()
    flash(t("patients.family_deleted").replace("{n}", str(len(freed))), "info")
    patient_id = request.form.get("patient_id", type=int)
    if patient_id:
        return redirect(url_for("patients.view", patient_id=patient_id) + "#family")
    return redirect(url_for("patients.index"))


@patients_bp.route("/<int:patient_id>/siblings/auto", methods=["POST"])
@module_required(MODULE)
def sibling_auto_link(patient_id):
    """Link the one candidate that matches on both signals.

    Offered as a press rather than done silently on save. The import already
    groups children by their father's name and gets it wrong on real data;
    doing more of that automatically, invisibly, is how a clinic ends up
    distrusting every family on the screen. One press, one child, and the link
    it makes says it was the program's.
    """
    from app.utils.siblings import auto_link

    patient = db.get_or_404(Patient, patient_id)
    if not patient.family_id:
        flash(t("siblings.need_family"), "warning")
        return redirect(url_for("patients.view", patient_id=patient.id) + "#family")

    other = auto_link(patient)
    if other is None:
        flash(t("siblings.no_certain_match"), "info")
    else:
        ActivityLog.record("patient.sibling.auto_link", user_id=current_user.id,
                           entity="patient", entity_id=patient.id,
                           detail=other.patient_number, ip_address=client_ip())
        db.session.commit()
        flash(t("siblings.auto_linked").replace(
            "{name}", other.display_name(getattr(g, "lang", "ar"))), "success")
    return redirect(url_for("patients.view", patient_id=patient.id) + "#family")
