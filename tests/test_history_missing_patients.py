"""What happens to a child in the services file who is not in the program yet.

Asked directly, and only half of it had an answer.

**Their history is not lost, and never was.** Nothing is imported for them and
nothing is discarded either — the rows stay in the clinic's own file, and a
re-upload once the patients exist picks them up, because a second upload
compares instead of appending. That was true and the screen never said it, so
somebody reading "412 rows rejected" had no way to know the work was still
recoverable.

**And "import those patients first" was a sentence, not a path.** A services
export carries no date of birth, gender or phone, so the patients cannot be
created from it — and telling a clinic to type eighty-seven children by hand
is how it decides to skip its own history instead.

So the file they need is produced for them: the patient-import sheet, with
the code and name already in it from their own export, and the columns only
they can supply left blank. **The date of birth is deliberately one of those.**
It drives every vaccination due date and every growth centile, so a guessed
one is a wrong dose date for a real child — worth more than the convenience of
not typing it.
"""
import io
import os
import sys
from datetime import date, datetime, time

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pytest  # noqa: E402

HEADERS = ["م", "تاريخ الخدمة", "كود المريض", "اسم المريض", "الخدمة", "السعر الإجمالي"]
MAPPING = {"col_source_row": "0", "col_service_date": "1",
           "col_patient_code": "2", "col_patient_name": "3",
           "col_service_name": "4", "col_price": "5"}


def _sheet(rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _row(serial, code, name, when=date(2024, 3, 1), what="كشف"):
    return [serial, datetime.combine(when, time()), code, name, what, 200]


@pytest.fixture()
def boss(clinic):
    return clinic["sign_in"]("boss")


def _token(body):
    import re

    found = re.search(r'name="token" value="([0-9a-f]+)"', body)
    assert found, "the screen did not carry a token forward"
    return found.group(1)


def _preview(boss, rows):
    """Upload and walk to the preview. Returns (body, token)."""
    reply = boss.post("/patients/import/history",
                      data={"file": (_sheet(rows), "e.xlsx")},
                      content_type="multipart/form-data")
    token = _token(reply.get_data(as_text=True))
    boss.post("/patients/import/history/link", data={"token": token, **MAPPING})
    out = boss.post("/patients/import/history/link",
                    data={"token": token, "confirm": "1"})
    body = out.get_data(as_text=True)
    return body, _token(body)


# ==================================================== nothing is lost =======
def test_an_unknown_patients_rows_are_refused_not_imported(boss, clinic):
    """A service attached to nobody is worse than one not yet imported."""
    from app.models import ImportedService

    _body, token = _preview(boss, [_row(1, "9999", "طفل مجهول")])
    boss.post("/patients/import/history/commit", data={"token": token},
              follow_redirects=True)
    with clinic["app"].app_context():
        assert ImportedService.query.count() == 0


def test_the_same_file_works_once_the_patient_exists(boss, clinic):
    """The reassurance the screen now states: re-uploading after importing the
    patients picks the rows up, because a second upload compares."""
    from app.models import ImportedService, Patient

    _preview(boss, [_row(1, "9999", "طفل مجهول")])

    with clinic["app"].app_context():
        clinic["db"].session.add(Patient(
            patient_number="PM-9999", reference_number="9999",
            full_name="طفل مجهول", gender="male",
            date_of_birth=date(2021, 5, 5), is_active=True))
        clinic["db"].session.commit()

    _body, token = _preview(boss, [_row(1, "9999", "طفل مجهول")])
    boss.post("/patients/import/history/commit", data={"token": token},
              follow_redirects=True)
    with clinic["app"].app_context():
        assert ImportedService.query.count() == 1


def test_the_screen_says_the_rows_are_kept(boss, clinic):
    """"412 rows rejected" with no more than that is how somebody concludes
    the history is unrecoverable and stops."""
    body, _token_ = _preview(boss, [_row(1, "9999", "طفل مجهول")])
    with clinic["app"].test_request_context("/"):
        from app.i18n import t
        assert t("history_import.missing_kept") in body


# ============================================== and the file to fill in =====
def _download(boss, rows):
    _body, token = _preview(boss, rows)
    return boss.post("/patients/import/history/missing", data={"token": token})


def _sheet_rows(reply):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(reply.data))
    ws = wb["Patients"]
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_the_missing_patients_come_back_as_a_patient_import_sheet(boss, clinic):
    from app.utils.imports import IMPORT_COLUMNS

    reply = _download(boss, [_row(1, "9999", "طفل مجهول")])
    assert reply.status_code == 200
    rows = _sheet_rows(reply)
    assert rows[0] == [c[0] for c in IMPORT_COLUMNS], "not the import's own columns"


def test_the_code_lands_in_the_column_the_history_matches_on(boss, clinic):
    """``reference_number`` is what the second upload looks the patient up by.
    Putting the old program's code anywhere else produces patients the
    re-upload still cannot find — which would make the whole round trip a
    waste of somebody's afternoon."""
    reply = _download(boss, [_row(1, "9999", "طفل مجهول")])
    rows = _sheet_rows(reply)
    headers, first = rows[0], rows[1]
    assert first[headers.index("reference_number")] == "9999"
    assert first[headers.index("full_name")] == "طفل مجهول"


def test_the_date_of_birth_is_left_blank_on_purpose(boss, clinic):
    """It decides every vaccination due date and every growth centile. A
    guessed one is a wrong dose date for a real child, and the clinic is the
    only place that knows it."""
    reply = _download(boss, [_row(1, "9999", "طفل مجهول")])
    rows = _sheet_rows(reply)
    headers, first = rows[0], rows[1]
    assert not first[headers.index("date_of_birth")]
    assert not first[headers.index("gender")]


def test_every_missing_patient_is_in_the_file_not_the_first_two_hundred(boss,
                                                                       clinic):
    """The screen is right to show the worst 200. A file somebody is going to
    fill in and upload has to be complete, or it silently loses the rest."""
    rows = [_row(i, str(10000 + i), f"طفل {i}") for i in range(1, 226)]
    reply = _download(boss, rows)
    assert len(_sheet_rows(reply)) == 226      # header + 225


def test_one_patient_with_many_services_is_one_row(boss, clinic):
    """It is a list of people, not of rows."""
    reply = _download(boss, [_row(1, "9999", "طفل مجهول"),
                             _row(2, "9999", "طفل مجهول"),
                             _row(3, "9999", "طفل مجهول")])
    assert len(_sheet_rows(reply)) == 2


def test_a_patient_who_does_exist_is_not_in_the_file(boss, clinic):
    """Offering to re-create somebody the clinic already has is how you get
    two files for one child."""
    from app.models import Patient

    with clinic["app"].app_context():
        clinic["db"].session.add(Patient(
            patient_number="PM-1001", reference_number="1001", full_name="معروف",
            gender="male", date_of_birth=date(2021, 1, 1), is_active=True))
        clinic["db"].session.commit()

    reply = _download(boss, [_row(1, "1001", "معروف"), _row(2, "9999", "مجهول")])
    codes = [r[0] for r in _sheet_rows(reply)[1:]]
    assert codes == ["9999"]


def test_a_file_with_nothing_missing_says_so_instead_of_sending_an_empty_sheet(
        boss, clinic):
    from app.models import Patient

    with clinic["app"].app_context():
        clinic["db"].session.add(Patient(
            patient_number="PM-1001", reference_number="1001", full_name="معروف",
            gender="male", date_of_birth=date(2021, 1, 1), is_active=True))
        clinic["db"].session.commit()

    reply = _download(boss, [_row(1, "1001", "معروف")])
    assert reply.status_code in (302, 303)


def test_the_sheet_explains_what_to_do_with_it(boss, clinic):
    """A spreadsheet of blanks with no instructions is a spreadsheet nobody
    fills in."""
    from openpyxl import load_workbook

    reply = _download(boss, [_row(1, "9999", "طفل مجهول")])
    wb = load_workbook(io.BytesIO(reply.data))
    assert "تعليمات" in wb.sheetnames
    text = " ".join(str(c.value or "") for row in wb["تعليمات"].iter_rows()
                    for c in row)
    assert "date_of_birth" in text


def test_the_button_is_on_the_preview(boss, clinic):
    body, _token_ = _preview(boss, [_row(1, "9999", "طفل مجهول")])
    assert "/patients/import/history/missing" in body


def test_an_expired_upload_does_not_produce_an_empty_file(boss, clinic):
    reply = boss.post("/patients/import/history/missing",
                      data={"token": "deadbeef"}, follow_redirects=True)
    with clinic["app"].test_request_context("/"):
        from app.i18n import t
        assert t("import.session_expired") in reply.get_data(as_text=True)


def test_both_languages_carry_the_new_words(clinic):
    import json

    root = os.path.join(os.path.dirname(__file__), "..")
    for lang in ("ar", "en"):
        with open(os.path.join(root, "app", "i18n", "locales", f"{lang}.json"),
                  encoding="utf-8") as fh:
            data = json.load(fh)
        for key in ("missing_kept", "missing_download", "no_missing",
                    "missing_sheet_1", "missing_sheet_2", "missing_sheet_3"):
            assert data["history_import"].get(key), f"{lang}.{key}"
