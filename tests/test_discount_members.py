"""A members list on every discount — the contract's screen, for a discount.

Eligibility here has always been *computed*: the club discount looks for a
card, the category discount reads the parents' category, the doctor's discount
looks at who is seeing the child. That is right for the common case and no use
at all for the two a clinic actually runs into, and both are lists:

* **"these families, by name"** — the doctors' children, the four families
  carried over from the old practice, with nothing on file a rule could match;
* **"everyone except him"** — one member whose discount has been stopped,
  without deleting the rule covering the other three hundred.

Two decisions hold the whole thing up.

**An exclusion always wins.** Taking somebody off a discount is an
instruction, and an instruction a rule can override is not one.

**And the list tops the rule up unless the clinic says otherwise.** Off by
default, so an existing discount gains a members list without one person's
eligibility changing. Assuming the other way — that naming one family silently
cuts off everybody else — would be a catastrophe nobody would notice for a
month.
"""
import os
import sys
from datetime import date

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pytest  # noqa: E402


@pytest.fixture()
def boss(clinic):
    return clinic["sign_in"]("boss")


@pytest.fixture()
def rule(clinic):
    """A staff-category discount, and a child whose parent is not staff."""
    from app.models import NamedDiscount

    with clinic["app"].app_context():
        row = NamedDiscount(name="خصم الموظفين", dtype="category", value=50,
                            is_percent=True, client_category="employee",
                            scope="all", is_active=True, auto_apply=True)
        clinic["db"].session.add(row)
        clinic["db"].session.commit()
        return row.id


def _applies(clinic, rule_id, patient_id=None):
    from app.models import NamedDiscount, Patient

    with clinic["app"].app_context():
        row = clinic["db"].session.get(NamedDiscount, rule_id)
        patient = clinic["db"].session.get(
            Patient, patient_id or clinic["ids"]["child"])
        return row.applies_to(patient=patient)


def _make_staff(clinic):
    """Put the fixture's child in a family whose father is staff."""
    from app.models import Family, Parent, Patient

    with clinic["app"].app_context():
        fam = Family(family_name="أسرة")
        clinic["db"].session.add(fam)
        clinic["db"].session.flush()
        clinic["db"].session.add(Parent(family_id=fam.id, full_name="أب",
                                        relation="father",
                                        client_category="employee"))
        child = clinic["db"].session.get(Patient, clinic["ids"]["child"])
        child.family_id = fam.id
        clinic["db"].session.commit()


def _add(boss, rule_id, patient_id, mode="include"):
    return boss.post(f"/finance/discounts/{rule_id}/members/add",
                     data={"patient_id": patient_id, "mode": mode},
                     follow_redirects=True)


# ================================================= naming somebody by hand ==
def test_the_rule_alone_does_not_reach_this_child(rule, clinic):
    """The starting point: no staff parent, so no discount."""
    assert _applies(clinic, rule) is False


def test_naming_them_gives_them_the_discount(boss, rule, clinic):
    """The case the computed rule cannot express — nothing on file to match."""
    _add(boss, rule, clinic["ids"]["child"])
    assert _applies(clinic, rule) is True


def test_naming_somebody_does_not_cut_everybody_else_off(boss, rule, clinic):
    """The default that makes this safe to switch on: the list tops the rule
    up. Assuming the other way would quietly stop three hundred discounts."""
    from app.models import Family, Parent, Patient

    with clinic["app"].app_context():
        fam = Family(family_name="أسرة موظف")
        clinic["db"].session.add(fam)
        clinic["db"].session.flush()
        clinic["db"].session.add(Parent(family_id=fam.id, full_name="أب",
                                        relation="father",
                                        client_category="employee"))
        other = Patient(patient_number="E1", full_name="ابن موظف",
                        gender="male", date_of_birth=date(2022, 1, 1),
                        family_id=fam.id, is_active=True)
        clinic["db"].session.add(other)
        clinic["db"].session.commit()
        other_id = other.id

    _add(boss, rule, clinic["ids"]["child"])
    assert _applies(clinic, rule, other_id) is True     # still covered by rule
    assert _applies(clinic, rule) is True               # and the named child


# ==================================================== taking somebody off ===
def test_excluding_somebody_the_rule_covers_stops_their_discount(boss, rule,
                                                                 clinic):
    """"Everyone except him", without deleting the rule."""
    _make_staff(clinic)
    assert _applies(clinic, rule) is True

    _add(boss, rule, clinic["ids"]["child"], mode="exclude")
    assert _applies(clinic, rule) is False


def test_an_exclusion_beats_being_named_as_well(boss, rule, clinic):
    """Adding then excluding the same person is a correction, and the last
    word is the exclusion — anything else makes "remove" advisory."""
    _add(boss, rule, clinic["ids"]["child"])
    _add(boss, rule, clinic["ids"]["child"], mode="exclude")
    assert _applies(clinic, rule) is False


def test_adding_the_same_person_twice_does_not_duplicate_them(boss, rule,
                                                              clinic):
    from app.models import DiscountMember

    _add(boss, rule, clinic["ids"]["child"])
    _add(boss, rule, clinic["ids"]["child"], mode="exclude")
    with clinic["app"].app_context():
        assert DiscountMember.query.filter_by(discount_id=rule).count() == 1


def test_removing_the_row_puts_them_back_under_the_rule(boss, rule, clinic):
    from app.models import DiscountMember

    _make_staff(clinic)
    _add(boss, rule, clinic["ids"]["child"], mode="exclude")
    assert _applies(clinic, rule) is False

    with clinic["app"].app_context():
        mid = DiscountMember.query.filter_by(discount_id=rule).one().id
    boss.post(f"/finance/discounts/members/{mid}/delete", follow_redirects=True)
    assert _applies(clinic, rule) is True


# ================================================ the list instead of the rule ==
def test_members_only_ignores_the_rule(boss, rule, clinic):
    """"The discount is for these families" — what a clinic means when it says
    that, and what would be a catastrophe to assume."""
    from app.models import NamedDiscount

    _make_staff(clinic)
    with clinic["app"].app_context():
        clinic["db"].session.get(NamedDiscount, rule).members_only = True
        clinic["db"].session.commit()

    assert _applies(clinic, rule) is False       # covered by the rule, not named
    _add(boss, rule, clinic["ids"]["child"])
    assert _applies(clinic, rule) is True


def test_members_only_with_an_empty_list_reaches_nobody(rule, clinic):
    """Rather than falling back to the rule, which would be the opposite of
    what the switch says."""
    from app.models import NamedDiscount

    with clinic["app"].app_context():
        clinic["db"].session.get(NamedDiscount, rule).members_only = True
        clinic["db"].session.commit()
    assert _applies(clinic, rule) is False


def test_it_is_off_by_default(rule, clinic):
    from app.models import NamedDiscount

    with clinic["app"].app_context():
        assert clinic["db"].session.get(NamedDiscount, rule).members_only is False


# ============================================================== the screen ==
def test_every_discount_has_a_members_screen(boss, rule, clinic):
    """Not just the clubs — a list is a list whatever the rule underneath is."""
    assert boss.get(f"/finance/discounts/{rule}/members").status_code == 200


def test_the_discounts_list_links_to_it(boss, rule, clinic):
    body = boss.get("/finance/discounts").get_data(as_text=True)
    assert f"/finance/discounts/{rule}/members" in body


def test_the_screen_shows_what_the_rule_alone_covers(boss, rule, clinic):
    """The one thing somebody misreads here is whether the list adds to the
    rule or replaces it, so both numbers are on the screen together."""
    body = boss.get(f"/finance/discounts/{rule}/members").get_data(as_text=True)
    with clinic["app"].test_request_context("/"):
        from app.i18n import t
        assert t("discount_members.rule_reach") in body


def test_the_screen_says_which_way_round_it_is(boss, rule, clinic):
    body = boss.get(f"/finance/discounts/{rule}/members").get_data(as_text=True)
    with clinic["app"].test_request_context("/"):
        from app.i18n import t
        assert t("discount_members.only_off") in body
        assert t("discount_members.exclusion_wins") in body


def test_the_list_comes_out_as_a_spreadsheet(boss, rule, clinic):
    """Asked for: a report of who is on a discount, for the accountant or the
    club."""
    import io

    from openpyxl import load_workbook

    _add(boss, rule, clinic["ids"]["child"])
    reply = boss.get(f"/finance/discounts/{rule}/members/export")
    assert reply.status_code == 200
    ws = load_workbook(io.BytesIO(reply.data)).active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert len(rows) == 2
    assert "طفل" in str(rows[1])


def test_adding_a_member_is_written_to_the_activity_log(boss, rule, clinic):
    """Somebody changed who gets money off. That is not a silent event."""
    from app.models import ActivityLog

    _add(boss, rule, clinic["ids"]["child"])
    with clinic["app"].app_context():
        assert ActivityLog.query.filter_by(action="discount.member.add").count() == 1


def test_a_patient_that_does_not_exist_is_refused(boss, rule, clinic):
    from app.models import DiscountMember

    _add(boss, rule, 999999)
    with clinic["app"].app_context():
        assert DiscountMember.query.count() == 0


def test_both_languages_carry_the_words(clinic):
    import json

    root = os.path.join(os.path.dirname(__file__), "..")
    for lang in ("ar", "en"):
        with open(os.path.join(root, "app", "i18n", "locales", f"{lang}.json"),
                  encoding="utf-8") as fh:
            data = json.load(fh)
        for key in ("title", "included", "excluded", "members_only",
                    "members_only_hint", "exclusion_wins", "rule_reach"):
            assert data["discount_members"].get(key), f"{lang}.{key}"
